"""
Comprehensive tests for the refactored cluster capacity report with modular design.

Tests cover:
- Multi-format support (HTML/Excel)
- Individual modular methods
- Data generation and processing
- Resource aggregation at namespace level
- Node capacity integration
- Edge cases and error handling
"""
import tempfile
import pytest
from pathlib import Path
from datetime import datetime, timezone

from data_gatherer.reporting.cluster_capacity_report import ClusterCapacityReport
from data_gatherer.persistence.db import WorkloadDB
from data_gatherer.util.hash import sha256_of_manifest


@pytest.fixture
def cluster_db_with_data(tmp_path):
    """Create a test database with sample cluster workload data."""
    db_path = tmp_path / 'cluster_test.db'
    db = WorkloadDB(str(db_path))
    now = datetime.now(timezone.utc)
    
    # Sample nodes for cluster capacity
    cur = db._conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS node_capacity (
            cluster TEXT, node_name TEXT, cpu_allocatable TEXT, memory_allocatable TEXT, 
            cpu_capacity TEXT, memory_capacity TEXT, node_role TEXT, deleted INTEGER DEFAULT 0, 
            first_seen TEXT, last_seen TEXT
        )
    """)
    
    # Insert sample nodes
    cur.execute("""
        INSERT INTO node_capacity (cluster, node_name, cpu_allocatable, memory_allocatable, 
                                  cpu_capacity, memory_capacity, node_role, deleted, first_seen, last_seen)
        VALUES 
        ('test-cluster', 'master-1', '3500m', '7168', '4000m', '8192', 'master', 0, '2025-01-01T00:00:00Z', '2025-01-01T00:00:00Z'),
        ('test-cluster', 'worker-1', '7500m', '14336', '8000m', '16384', 'worker', 0, '2025-01-01T00:00:00Z', '2025-01-01T00:00:00Z'),
        ('test-cluster', 'worker-2', '7500m', '14336', '8000m', '16384', 'worker', 0, '2025-01-01T00:00:00Z', '2025-01-01T00:00:00Z')
    """)
    
    # Sample workloads across multiple namespaces
    production_deployment = {
        'apiVersion': 'apps/v1',
        'kind': 'Deployment',
        'metadata': {'name': 'web-frontend', 'namespace': 'production'},
        'spec': {
            'replicas': 3,
            'template': {
                'spec': {
                    'containers': [
                        {
                            'name': 'web',
                            'image': 'nginx',
                            'resources': {
                                'requests': {'cpu': '200m', 'memory': '256Mi'},
                                'limits': {'cpu': '500m', 'memory': '512Mi'}
                            }
                        }
                    ]
                }
            }
        }
    }
    
    production_statefulset = {
        'apiVersion': 'apps/v1',
        'kind': 'StatefulSet',
        'metadata': {'name': 'database', 'namespace': 'production'},
        'spec': {
            'replicas': 2,
            'template': {
                'spec': {
                    'containers': [
                        {
                            'name': 'db',
                            'image': 'postgres',
                            'resources': {
                                'requests': {'cpu': '1000m', 'memory': '2Gi'},
                                'limits': {'cpu': '2000m', 'memory': '4Gi'}
                            }
                        }
                    ]
                }
            }
        }
    }
    
    development_deployment = {
        'apiVersion': 'apps/v1',
        'kind': 'Deployment',
        'metadata': {'name': 'test-app', 'namespace': 'development'},
        'spec': {
            'replicas': 1,
            'template': {
                'spec': {
                    'containers': [
                        {
                            'name': 'app',
                            'image': 'test-app',
                            'resources': {
                                'requests': {'cpu': '100m', 'memory': '128Mi'},
                                'limits': {'cpu': '200m', 'memory': '256Mi'}
                            }
                        }
                    ]
                }
            }
        }
    }
    
    # Insert workloads
    db.upsert_workload(
        cluster='test-cluster', api_version='apps/v1', kind='Deployment',
        namespace='production', name='web-frontend', resource_version='1', uid='u1',
        manifest=production_deployment, manifest_hash=sha256_of_manifest(production_deployment), now=now
    )
    
    db.upsert_workload(
        cluster='test-cluster', api_version='apps/v1', kind='StatefulSet',
        namespace='production', name='database', resource_version='1', uid='u2',
        manifest=production_statefulset, manifest_hash=sha256_of_manifest(production_statefulset), now=now
    )
    
    db.upsert_workload(
        cluster='test-cluster', api_version='apps/v1', kind='Deployment',
        namespace='development', name='test-app', resource_version='1', uid='u3',
        manifest=development_deployment, manifest_hash=sha256_of_manifest(development_deployment), now=now
    )
    
    db._conn.commit()
    return db, str(db_path)


@pytest.fixture
def empty_cluster_db(tmp_path):
    """Create an empty test database with node capacity but no workloads."""
    db_path = tmp_path / 'empty_cluster.db'
    db = WorkloadDB(str(db_path))
    
    # Add nodes but no workloads
    cur = db._conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS node_capacity (
            cluster TEXT, node_name TEXT, cpu_allocatable TEXT, memory_allocatable TEXT, 
            cpu_capacity TEXT, memory_capacity TEXT, node_role TEXT, deleted INTEGER DEFAULT 0, 
            first_seen TEXT, last_seen TEXT
        )
    """)
    
    cur.execute("""
        INSERT INTO node_capacity (cluster, node_name, cpu_allocatable, memory_allocatable, 
                                  cpu_capacity, memory_capacity, node_role, deleted, first_seen, last_seen)
        VALUES ('test-cluster', 'master-1', '4000m', '8192', '4000m', '8192', 'master', 0, '2025-01-01T00:00:00Z', '2025-01-01T00:00:00Z')
    """)
    
    db._conn.commit()
    return db, str(db_path)


class TestClusterCapacityReportDataGeneration:
    """Test the core data generation functionality."""
    
    def test_generate_capacity_data_with_workloads(self, cluster_db_with_data):
        """Test data generation with sample workloads across namespaces."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        
        # Verify structure
        assert 'ns_totals' in capacity_data
        assert 'node_capacity' in capacity_data
        assert 'summary_totals' in capacity_data
        
        # Verify namespace data
        ns_totals = capacity_data['ns_totals']
        assert len(ns_totals) == 2  # production and development
        assert 'production' in ns_totals
        assert 'development' in ns_totals
        
        # Verify production namespace totals (3 web replicas + 2 db replicas)
        prod_ns = ns_totals['production']
        assert prod_ns['cpu'] == 2600  # (200*3) + (1000*2) = 600 + 2000
        assert prod_ns['mem'] == 4864  # (256*3) + (2048*2) = 768 + 4096
        assert prod_ns['cpu_lim'] == 5500  # (500*3) + (2000*2) = 1500 + 4000
        assert prod_ns['mem_lim'] == 9728  # (512*3) + (4096*2) = 1536 + 8192
        
        # Verify development namespace totals (1 replica)
        dev_ns = ns_totals['development']
        assert dev_ns['cpu'] == 100
        assert dev_ns['mem'] == 128
        assert dev_ns['cpu_lim'] == 200
        assert dev_ns['mem_lim'] == 256
        
        # Verify node capacity (only worker nodes counted)
        node_capacity = capacity_data['node_capacity']
        assert node_capacity['total_cpu_alloc'] == 15000  # 7500m + 7500m (workers only)
        assert node_capacity['total_mem_alloc'] == 28672  # 14336 + 14336 (workers only)
        
        # Verify summary totals
        summary = capacity_data['summary_totals']
        assert summary['total_req_cpu'] == 2700  # 2600 + 100
        assert summary['total_req_mem'] == 4992  # 4864 + 128
        assert summary['total_lim_cpu'] == 5700  # 5500 + 200
        assert summary['total_lim_mem'] == 9984  # 9728 + 256
    
    def test_generate_capacity_data_empty_cluster(self, empty_cluster_db):
        """Test data generation with empty cluster (no workloads)."""
        db, _ = empty_cluster_db
        report = ClusterCapacityReport()
        
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        
        # Verify structure exists
        assert 'ns_totals' in capacity_data
        assert 'node_capacity' in capacity_data
        assert 'summary_totals' in capacity_data
        
        # Should have empty namespaces
        assert capacity_data['ns_totals'] == {}
        
        # Should still have node capacity (master is not counted as it's not a worker)
        node_capacity = capacity_data['node_capacity']
        assert node_capacity['total_cpu_alloc'] == 0  # No worker nodes
        assert node_capacity['total_mem_alloc'] == 0  # No worker nodes
        
        # Summary should be zeros
        summary = capacity_data['summary_totals']
        assert summary['total_req_cpu'] == 0
        assert summary['total_req_mem'] == 0
        assert summary['total_lim_cpu'] == 0
        assert summary['total_lim_mem'] == 0


class TestClusterCapacityReportModularMethods:
    """Test individual modular methods."""
    
    def test_process_namespace_totals(self, cluster_db_with_data):
        """Test namespace totals processing."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        ns_totals = report._process_namespace_totals(db, 'test-cluster')
        
        # Should have two namespaces
        assert len(ns_totals) == 2
        assert 'production' in ns_totals
        assert 'development' in ns_totals
        
        # Verify production totals
        prod = ns_totals['production']
        assert prod['cpu'] == 2600
        assert prod['mem'] == 4864
        assert prod['cpu_lim'] == 5500
        assert prod['mem_lim'] == 9728
        
        # Verify development totals
        dev = ns_totals['development']
        assert dev['cpu'] == 100
        assert dev['mem'] == 128
        assert dev['cpu_lim'] == 200
        assert dev['mem_lim'] == 256
    
    def test_process_container_resources(self, cluster_db_with_data):
        """Test container resource processing logic."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        # Test with a container definition
        cdef = {
            'name': 'test-container',
            'resources': {
                'requests': {'cpu': '200m', 'memory': '256Mi'},
                'limits': {'cpu': '400m', 'memory': '512Mi'}
            }
        }
        
        ns_total = {'cpu': 0, 'mem': 0, 'cpu_lim': 0, 'mem_lim': 0}
        
        # Process 3 replicas
        report._process_container_resources(cdef, 3, ns_total)
        
        # Verify accumulation
        assert ns_total['cpu'] == 600  # 200 * 3
        assert ns_total['mem'] == 768  # 256 * 3
        assert ns_total['cpu_lim'] == 1200  # 400 * 3
        assert ns_total['mem_lim'] == 1536  # 512 * 3
    
    def test_process_container_resources_missing_values(self, cluster_db_with_data):
        """Test container resource processing with missing resources."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        # Test with missing limits
        cdef = {
            'name': 'test-container',
            'resources': {
                'requests': {'cpu': '100m', 'memory': '128Mi'}
                # No limits
            }
        }
        
        ns_total = {'cpu': 0, 'mem': 0, 'cpu_lim': 0, 'mem_lim': 0}
        
        report._process_container_resources(cdef, 2, ns_total)
        
        # Should accumulate requests but not limits
        assert ns_total['cpu'] == 200
        assert ns_total['mem'] == 256
        assert ns_total['cpu_lim'] == 0
        assert ns_total['mem_lim'] == 0
    
    def test_get_node_capacity(self, cluster_db_with_data):
        """Test node capacity aggregation."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        node_capacity = report._get_node_capacity(db, 'test-cluster')
        
        # Should sum worker nodes only
        assert node_capacity['total_cpu_alloc'] == 15000  # 7500m + 7500m (workers only)
        assert node_capacity['total_mem_alloc'] == 28672  # 14336 + 14336 (workers only)
    
    def test_get_node_capacity_no_nodes(self, tmp_path):
        """Test node capacity with no nodes."""
        db_path = tmp_path / 'no_nodes.db'
        db = WorkloadDB(str(db_path))
        report = ClusterCapacityReport()
        
        node_capacity = report._get_node_capacity(db, 'test-cluster')
        
        # Should return zeros
        assert node_capacity['total_cpu_alloc'] == 0
        assert node_capacity['total_mem_alloc'] == 0
    
    def test_calculate_summary_totals(self, cluster_db_with_data):
        """Test summary totals calculation."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        # Mock namespace totals
        ns_totals = {
            'production': {
                'cpu': 1000, 'mem': 2048,
                'cpu_lim': 2000, 'mem_lim': 4096
            },
            'development': {
                'cpu': 500, 'mem': 1024,
                'cpu_lim': 1000, 'mem_lim': 2048
            }
        }
        
        summary = report._calculate_summary_totals(ns_totals)
        
        assert summary['total_req_cpu'] == 1500
        assert summary['total_req_mem'] == 3072
        assert summary['total_lim_cpu'] == 3000
        assert summary['total_lim_mem'] == 6144


class TestClusterCapacityReportHTMLGeneration:
    """Test HTML report generation."""
    
    def test_generate_html_report_with_data(self, cluster_db_with_data, tmp_path):
        """Test HTML report generation with workload data."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        html_content = report._generate_html_report('Test Cluster Capacity', capacity_data, 'test-cluster')
        
        # Verify HTML structure
        assert '<html>' in html_content
        assert 'Test Cluster Capacity' in html_content
        assert 'production' in html_content
        assert 'development' in html_content
        assert 'Totals' in html_content
        
        # Verify CSS is included
        assert '<style>' in html_content
        assert 'table' in html_content
        assert 'background-color' in html_content
        
        # Verify data values are present (formatted with commas in HTML)
        assert '2,600' in html_content  # production CPU requests
        assert '100' in html_content   # development CPU requests
    
    def test_generate_html_report_empty(self, empty_cluster_db, tmp_path):
        """Test HTML report generation with empty cluster."""
        db, _ = empty_cluster_db
        report = ClusterCapacityReport()
        
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        html_content = report._generate_html_report('Empty Cluster', capacity_data, 'test-cluster')
        
        # Should still generate valid HTML
        assert '<html>' in html_content
        assert 'Empty Cluster' in html_content
        assert 'No workloads found' in html_content or 'Totals' in html_content
    
    def test_full_html_generation(self, cluster_db_with_data, tmp_path):
        """Test full HTML report generation via generate method."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        html_path = tmp_path / 'cluster_capacity.html'
        report.generate(db, 'test-cluster', str(html_path), format='html')
        
        assert html_path.exists()
        
        # Read and verify content
        html_content = html_path.read_text()
        assert '<html>' in html_content
        assert 'production' in html_content
        assert 'development' in html_content


class TestClusterCapacityReportExcelGeneration:
    """Test Excel report generation."""
    
    def test_generate_excel_report_with_data(self, cluster_db_with_data, tmp_path):
        """Test Excel report generation with workload data."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        excel_path = tmp_path / 'cluster_capacity.xlsx'
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        report._generate_excel_report('Test Cluster Capacity', capacity_data, str(excel_path))
        
        assert excel_path.exists()
        
        # Verify Excel content using openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(str(excel_path))
        ws = wb.active
        
        # Check title
        assert ws.title == "Cluster Capacity Report"
        
        # Check that namespace data is present
        found_production = False
        found_development = False
        for row in ws.iter_rows():
            row_values = [str(cell.value) if cell.value is not None else '' for cell in row]
            if 'production' in row_values:
                found_production = True
            if 'development' in row_values:
                found_development = True
        
        assert found_production
        assert found_development
        wb.close()
    
    def test_generate_excel_report_empty(self, empty_cluster_db, tmp_path):
        """Test Excel report generation with empty cluster."""
        db, _ = empty_cluster_db
        report = ClusterCapacityReport()
        
        excel_path = tmp_path / 'empty_cluster_capacity.xlsx'
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        report._generate_excel_report('Empty Cluster', capacity_data, str(excel_path))
        
        assert excel_path.exists()
        
        # Verify basic structure
        from openpyxl import load_workbook
        wb = load_workbook(str(excel_path))
        ws = wb.active
        assert ws.title == "Cluster Capacity Report"
        wb.close()
    
    def test_excel_basic_formatting(self, cluster_db_with_data, tmp_path):
        """Test Excel basic formatting and structure."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        excel_path = tmp_path / 'cluster_capacity_formatted.xlsx'
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        report._generate_excel_report('Test Cluster', capacity_data, str(excel_path))
        
        assert excel_path.exists()
        
        # Verify Excel structure is valid
        from openpyxl import load_workbook
        wb = load_workbook(str(excel_path))
        ws = wb.active
        
        # Check that we have data rows
        assert ws.max_row > 5  # Should have headers and data
        assert ws.max_column >= 7  # Should have all columns
        wb.close()
    
    def test_excel_openpyxl_import_error(self, cluster_db_with_data, tmp_path, monkeypatch):
        """Test Excel generation when openpyxl is not available."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        # Mock import error
        import builtins
        original_import = builtins.__import__
        
        def mock_import(name, *args, **kwargs):
            if name == 'openpyxl':
                raise ImportError("openpyxl not available")
            return original_import(name, *args, **kwargs)
        
        monkeypatch.setattr(builtins, '__import__', mock_import)
        
        excel_path = tmp_path / 'cluster_capacity_no_openpyxl.xlsx'
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        
        # Should raise ImportError
        with pytest.raises(ImportError, match="openpyxl is required for Excel output"):
            report._generate_excel_report('Test Cluster', capacity_data, str(excel_path))


class TestClusterCapacityReportEdgeCases:
    """Test edge cases and error handling."""
    
    def test_resource_parsing_edge_cases(self, cluster_db_with_data):
        """Test parsing of various resource formats."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        # Test various CPU formats
        test_cases = [
            ('100m', 100),
            ('1', 1000),
            ('1.5', 1500),
            ('2000m', 2000),
            ('0.1', 100),
            ('', 0),
            (None, 0),
        ]
        
        for cpu_value, expected in test_cases:
            cdef = {
                'name': 'test',
                'resources': {
                    'requests': {'cpu': cpu_value} if cpu_value is not None else {}
                }
            }
            ns_total = {'cpu': 0, 'mem': 0, 'cpu_lim': 0, 'mem_lim': 0}
            
            report._process_container_resources(cdef, 1, ns_total)
            assert ns_total['cpu'] == expected, f"Failed for CPU value: {cpu_value}"
    
    def test_memory_parsing_edge_cases(self, cluster_db_with_data):
        """Test parsing of various memory formats."""
        db, _ = cluster_db_with_data
        report = ClusterCapacityReport()
        
        # Test various memory formats
        test_cases = [
            ('128Mi', 128),
            ('1Gi', 1024),
            ('512Mi', 512),
            ('2Gi', 2048),
            ('1024Mi', 1024),
            ('', 0),
            (None, 0),
        ]
        
        for memory_value, expected in test_cases:
            cdef = {
                'name': 'test',
                'resources': {
                    'requests': {'memory': memory_value} if memory_value is not None else {}
                }
            }
            ns_total = {'cpu': 0, 'mem': 0, 'cpu_lim': 0, 'mem_lim': 0}
            
            report._process_container_resources(cdef, 1, ns_total)
            assert ns_total['mem'] == expected, f"Failed for memory value: {memory_value}"
    
    def test_invalid_manifest_handling(self, tmp_path):
        """Test handling of invalid manifests."""
        db_path = tmp_path / 'invalid_manifest.db'
        db = WorkloadDB(str(db_path))
        
        # Insert invalid manifest
        cur = db._conn.cursor()
        cur.execute("""
            INSERT INTO workload (cluster, kind, namespace, name, api_version, manifest_json, 
                                 manifest_hash, deleted, first_seen, last_seen)
            VALUES ('test-cluster', 'Deployment', 'test', 'invalid', 'apps/v1', 
                   'invalid-json', 'hash', 0, '2025-01-01T00:00:00Z', '2025-01-01T00:00:00Z')
        """)
        db._conn.commit()
        
        report = ClusterCapacityReport()
        
        # Should handle gracefully without crashing
        capacity_data = report._generate_capacity_data(db, 'test-cluster')
        
        # Should still return valid structure
        assert 'ns_totals' in capacity_data
        assert 'node_capacity' in capacity_data
        assert 'summary_totals' in capacity_data


def test_cluster_capacity_report_cli_integration(tmp_path):
    """Test full CLI integration with cluster capacity report."""
    db_path = tmp_path / 'cli_test.db'
    db = WorkloadDB(str(db_path))
    now = datetime.now(timezone.utc)
    
    # Insert minimal data
    cur = db._conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS node_capacity (
            cluster TEXT, node_name TEXT, cpu_allocatable TEXT, memory_allocatable TEXT, 
            cpu_capacity TEXT, memory_capacity TEXT, node_role TEXT, deleted INTEGER DEFAULT 0, 
            first_seen TEXT, last_seen TEXT
        )
    """)
    cur.execute("""
        INSERT INTO node_capacity (cluster, node_name, cpu_allocatable, memory_allocatable, 
                                  cpu_capacity, memory_capacity, node_role, deleted, first_seen, last_seen)
        VALUES ('cli-cluster', 'node1', '4000', '8192', '4000', '8192', 'worker', 0, 
                '2025-01-01T00:00:00Z', '2025-01-01T00:00:00Z')
    """)
    
    # Insert simple workload
    manifest = {
        'apiVersion': 'apps/v1',
        'kind': 'Deployment',
        'metadata': {'name': 'cli-app', 'namespace': 'default'},
        'spec': {
            'replicas': 1,
            'template': {
                'spec': {
                    'containers': [
                        {
                            'name': 'app',
                            'resources': {
                                'requests': {'cpu': '100m', 'memory': '128Mi'},
                                'limits': {'cpu': '200m', 'memory': '256Mi'}
                            }
                        }
                    ]
                }
            }
        }
    }
    
    db.upsert_workload(
        cluster='cli-cluster', api_version='apps/v1', kind='Deployment',
        namespace='default', name='cli-app', resource_version='1', uid='cli-uid',
        manifest=manifest, manifest_hash=sha256_of_manifest(manifest), now=now
    )
    
    # Test both formats
    html_path = tmp_path / 'cli_cluster.html'
    excel_path = tmp_path / 'cli_cluster.xlsx'
    
    report = ClusterCapacityReport()
    
    # Generate HTML
    report.generate(db, 'cli-cluster', str(html_path), format='html')
    assert html_path.exists()
    assert 'default' in html_path.read_text()
    
    # Generate Excel
    report.generate(db, 'cli-cluster', str(excel_path), format='excel')
    assert excel_path.exists()
