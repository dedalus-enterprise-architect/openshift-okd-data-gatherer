"""New tests for ClusterCapacityReport aligned with current implementation.

Focus areas:
1. Core data structure (ns_totals, node_capacity, summary_totals, ns_details)
2. HTML generation sections
3. Excel generation structural validation
4. CLI integration (cluster-capacity type)
"""
from datetime import datetime, timezone
import pytest
from data_gatherer.reporting.cluster_capacity_report import ClusterCapacityReport
from data_gatherer.persistence.db import WorkloadDB
from data_gatherer.util.hash import sha256_of_manifest


@pytest.fixture
def db_with_sample(tmp_path):
    db_path = tmp_path / 'sample.db'
    db = WorkloadDB(str(db_path))
    now = datetime.now(timezone.utc)
    # Two workloads across two namespaces
    man1 = {
        'apiVersion': 'apps/v1', 'kind': 'Deployment',
        'metadata': {'name': 'app', 'namespace': 'ns1'},
        'spec': {'replicas': 2, 'template': {'spec': {'containers': [
            {'name': 'c1', 'resources': {'requests': {'cpu': '250m', 'memory': '256Mi'}, 'limits': {'cpu': '500m', 'memory': '512Mi'}}},
            {'name': 'c2', 'resources': {'requests': {'cpu': '100m', 'memory': '64Mi'}}}
        ]}}}
    }
    man2 = {
        'apiVersion': 'apps/v1', 'kind': 'StatefulSet',
        'metadata': {'name': 'db', 'namespace': 'ns2'},
        'spec': {'replicas': 1, 'template': {'spec': {'containers': [
            {'name': 'pg', 'resources': {'requests': {'cpu': '300m', 'memory': '512Mi'}, 'limits': {'cpu': '600m', 'memory': '1Gi'}}}
        ]}}}
    }
    db.upsert_workload('clusterA', 'apps/v1', 'Deployment', 'ns1', 'app', '1', 'u1', man1, sha256_of_manifest(man1), now)
    db.upsert_workload('clusterA', 'apps/v1', 'StatefulSet', 'ns2', 'db', '1', 'u2', man2, sha256_of_manifest(man2), now)
    return db


def test_generate_capacity_data_structure(db_with_sample):
    report = ClusterCapacityReport()
    data = report._generate_capacity_data(db_with_sample, 'clusterA')
    for key in ['ns_totals', 'node_capacity', 'summary_totals', 'ns_details']:
        assert key in data
    # ns_totals correctness (init containers ignored, only main containers as implemented)
    # ns1: (c1 250m + c2 100m)=350m *2 replicas = 700m CPU, (256+64)=320Mi *2 = 640Mi
    # limits: c1 only 500m *2=1000m CPU, 512Mi*2=1024Mi
    assert data['ns_totals']['ns1']['cpu'] == 700
    assert data['ns_totals']['ns1']['mem'] == 640
    assert data['ns_totals']['ns1']['cpu_lim'] == 1000
    assert data['ns_totals']['ns1']['mem_lim'] == 1024
    # ns2: single replica 300m / 512Mi requests; limits 600m / 1024Mi
    assert data['ns_totals']['ns2']['cpu'] == 300
    assert data['ns_totals']['ns2']['mem'] == 512
    assert data['ns_totals']['ns2']['cpu_lim'] == 600
    assert data['ns_totals']['ns2']['mem_lim'] == 1024
    # Summary totals
    st = data['summary_totals']
    assert st['total_req_cpu'] == 1000  # 700 + 300
    assert st['total_req_mem'] == 1152  # 640 + 512
    assert st['total_lim_cpu'] == 1600  # 1000 + 600
    assert st['total_lim_mem'] == 2048  # 1024 + 1024
    # ns_details contains per-namespace lists
    assert 'ns1' in data['ns_details'] and 'ns2' in data['ns_details']
    assert len(data['ns_details']['ns1']) == 2
    assert len(data['ns_details']['ns2']) == 1


def test_generate_capacity_data_empty(tmp_path):
    db = WorkloadDB(str(tmp_path / 'empty.db'))
    report = ClusterCapacityReport()
    data = report._generate_capacity_data(db, 'empty')
    assert data['ns_totals'] == {}
    assert data['summary_totals']['total_req_cpu'] == 0
    assert data['ns_details'] == {}


def test_html_report_generation(db_with_sample, tmp_path):
    report = ClusterCapacityReport()
    data = report._generate_capacity_data(db_with_sample, 'clusterA')
    html_doc = report._generate_html_report('Cluster Capacity Report: clusterA', data, 'clusterA')
    # Core sections
    assert 'Container Requests vs Allocatable resources on Worker Nodes' in html_doc
    assert 'Namespace capacity vs Cluster capacity' in html_doc
    # Namespace names and totals row label
    assert 'ns1' in html_doc and 'ns2' in html_doc
    assert '<strong>Totals</strong>' in html_doc


def test_excel_report_generation(db_with_sample, tmp_path):
    pytest.importorskip('openpyxl')
    report = ClusterCapacityReport()
    out = tmp_path / 'cap.xlsx'
    report.generate(db_with_sample, 'clusterA', str(out), 'excel')
    assert out.exists()
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    ws = wb.active
    # Title
    assert 'Cluster Capacity Report: clusterA' in str(ws['A1'].value)
    # Find summary header row containing 'Scope'
    scope_row = None
    for r in range(1, 30):
        if ws.cell(row=r, column=1).value == 'Scope':
            scope_row = r
            break
    assert scope_row, 'Scope header row not found'
    expected_summary_headers = ['Scope', 'CPU (m)', 'CPU % Allocatable', 'Memory (Mi)', 'Memory % Allocatable']
    assert [ws.cell(row=scope_row, column=i).value for i in range(1, 6)] == expected_summary_headers
    # Find namespace header row containing 'Namespace'
    ns_header_row = None
    for r in range(scope_row + 1, scope_row + 50):
        if ws.cell(row=r, column=1).value == 'Namespace':
            ns_header_row = r
            break
    assert ns_header_row, 'Namespace header row not found'
    expected_ns_headers = [
        'Namespace', 'CPU Requests (m)', 'Memory Requests (Mi)',
        'CPU Limits (m)', 'Memory Limits (Mi)', '% CPU allocated on Cluster', '% Memory allocated on Cluster'
    ]
    assert [ws.cell(row=ns_header_row, column=i).value for i in range(1, 8)] == expected_ns_headers
    # Totals row after namespace rows
    totals_row_found = False
    for r in range(ns_header_row + 1, ns_header_row + 40):
        if ws.cell(row=r, column=1).value == 'Totals':
            totals_row_found = True
            break
    assert totals_row_found, 'Totals row not found in namespace capacity section'
    # Namespace detail section headers (Kind, Workload Name ...)
    detail_header_found = False
    for r in range(ns_header_row + 1, ns_header_row + 200):
        if ws.cell(row=r, column=1).value == 'Kind' and ws.cell(row=r, column=2).value == 'Workload Name':
            detail_header_found = True
            break
    assert detail_header_found, 'Detail header row not found'


def test_cli_integration_cluster_capacity(tmp_path):
    from click.testing import CliRunner
    from data_gatherer.run import cli
    runner = CliRunner()
    cfg = tmp_path / 'cfg.yaml'
    cfg.write_text(f"""storage:\n  base_dir: {tmp_path.as_posix()}\nclusters:\n  - name: c1\n    credentials:\n      host: https://dummy\n      verify_ssl: false\n    include_kinds: [Deployment]\n    parallelism: 1\nlogging:\n  level: INFO\n  format: text\n""")
    # init storage
    res = runner.invoke(cli, ['--config', str(cfg), 'init', '--cluster', 'c1'])
    assert res.exit_code == 0, res.output
    # Insert minimal workload directly
    db_path = tmp_path / 'c1' / 'data.db'
    db = WorkloadDB(str(db_path))
    now = datetime.now(timezone.utc)
    manifest = {
        'apiVersion': 'apps/v1', 'kind': 'Deployment',
        'metadata': {'name': 'demo', 'namespace': 'demo-ns'},
        'spec': {'replicas': 1, 'template': {'spec': {'containers': [
            {'name': 'c', 'resources': {'requests': {'cpu': '100m', 'memory': '128Mi'}}}
        ]}}}
    }
    db.upsert_workload('c1', 'apps/v1', 'Deployment', 'demo-ns', 'demo', '1', 'u1', manifest, sha256_of_manifest(manifest), now)
    db._conn.close()
    res2 = runner.invoke(cli, ['--config', str(cfg), 'report', '--cluster', 'c1', '--type', 'cluster-capacity'])
    assert res2.exit_code == 0, res2.output
    report_dir = tmp_path / 'c1' / 'reports'
    html_files = list(report_dir.glob('cluster-capacity-*.html'))
    assert html_files, 'No cluster capacity report generated'
    content = html_files[0].read_text()
    assert 'demo-ns' in content and 'Cluster Capacity Report: c1' in content
