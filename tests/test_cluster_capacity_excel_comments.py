"""
Test that container capacity report adds proper comments to namespace totals in Excel format.
"""
import pytest
import tempfile
import os
from datetime import datetime, timezone
from data_gatherer.reporting.cluster_capacity_report import ClusterCapacityReport
from data_gatherer.persistence.db import WorkloadDB
from data_gatherer.util.hash import sha256_of_manifest


def test_excel_namespace_totals_with_comments(sample_db):
    """Test that namespace totals cells have descriptive formula comments."""
    # Setup report generator and temporary output
    report = CapacityReport()
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "test-container-capacity.xlsx")
        
        # Generate Excel report
        report.generate(sample_db, "test-cluster", out_path, format='excel')
        
        # Verify file was created
        assert os.path.exists(out_path)
        
        # Load workbook and check for comments
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        
        wb = load_workbook(out_path)
        ws = wb.active
        
        # Find namespace totals rows (they should have "Namespace totals" in column 1)
        ns_totals_rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=False), start=4):
            if row[0].value == "Namespace totals":
                ns_totals_rows.append(row_idx)
        
        # Verify we found at least one namespace totals row
        assert len(ns_totals_rows) > 0, "Expected to find namespace totals rows"
        
        # Check that ALL metric cells (columns 7-14) have comments
        expected_columns = [7, 8, 9, 10, 11, 12, 13, 14]
        column_names = ['CPU_req_m', 'CPU_lim_m', 'Mem_req_Mi', 'Mem_lim_Mi', 
                       'CPU_req_m_total', 'CPU_lim_m_total', 'Mem_req_Mi_total', 'Mem_lim_Mi_total']
        
        for row_idx in ns_totals_rows:
            for col_num, col_name in zip(expected_columns, column_names):
                cell = ws.cell(row=row_idx, column=col_num)
                assert cell.comment is not None, \
                    f"Expected comment on namespace totals {col_name} cell at row {row_idx}, column {col_num}"
                
                # Verify comment contains expected text
                comment_text = cell.comment.text
                assert "namespace" in comment_text.lower(), \
                    f"Comment should mention 'namespace' for {col_name}"
                assert "Formula:" in comment_text or "formula" in comment_text.lower(), \
                    f"Comment should describe the formula for {col_name}"
                # No longer require init exclusion or runtime wording
                # Verify specific content based on column type
                if "_total" in col_name:
                    assert "Replicas" in comment_text or "×" in comment_text, \
                        f"Total columns should mention replica multiplication for {col_name}"


def test_excel_namespace_totals_values_match_html(sample_db):
    """Test that namespace totals in Excel match those in HTML report."""
    report = CapacityReport()
    
    # Generate both formats
    with tempfile.TemporaryDirectory() as tmpdir:
        html_path = os.path.join(tmpdir, "test.html")
        excel_path = os.path.join(tmpdir, "test.xlsx")
        
        report.generate(sample_db, "test-cluster", html_path, format='html')
        report.generate(sample_db, "test-cluster", excel_path, format='excel')
        
        assert os.path.exists(html_path)
        assert os.path.exists(excel_path)
        
        # Extract namespace totals from HTML
        with open(html_path, 'r') as f:
            html_content = f.read()
        
        # Load Excel and extract namespace totals
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Find namespace totals in Excel
        excel_ns_totals = {}
        current_namespace = None
        for row in ws.iter_rows(min_row=4, values_only=True):
            # Track current namespace - namespace header has value in col 0 but rest are None (merged cells)
            if row[0] and row[0] != "Namespace totals" and row[0] not in ["Totals (main containers)"]:
                # Check if this is a namespace header (merged row) or a workload row
                # Namespace header: has namespace in col 0, but columns 1-13 are None
                # Workload row: has Kind in col 0, Namespace in col 1
                if row[1] is None and row[2] is None:  # Likely a namespace header
                    current_namespace = row[0]
            elif row[0] == "Namespace totals" and current_namespace:
                # Extract totals (columns 7-14 correspond to indices 6-13)
                excel_ns_totals[current_namespace] = {
                    'cpu_req': row[6],
                    'cpu_lim': row[7],
                    'mem_req': row[8],
                    'mem_lim': row[9],
                    'cpu_req_total': row[10],
                    'cpu_lim_total': row[11],
                    'mem_req_total': row[12],
                    'mem_lim_total': row[13]
                }
        
        # Verify we found namespace totals in Excel
        assert len(excel_ns_totals) > 0, "Expected to find namespace totals in Excel"
        
        # Verify namespace totals appear in HTML
        for ns, totals in excel_ns_totals.items():
            assert f"<strong>{ns}</strong>" in html_content or ns in html_content, \
                f"Namespace {ns} should appear in HTML report"
            # Check that totals values appear in HTML (allowing for formatting differences)
            assert str(totals['cpu_req_total']) in html_content, \
                f"CPU request total {totals['cpu_req_total']} should appear in HTML"


def test_excel_namespace_grouping(sample_db):
    """Test that workloads are properly grouped by namespace in Excel."""
    report = CapacityReport()
    
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "test.xlsx")
        report.generate(sample_db, "test-cluster", out_path, format='excel')
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        
        wb = load_workbook(out_path)
        ws = wb.active
        
        # Track namespace order and grouping
        namespaces_seen = []
        current_namespace = None
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            # Namespace header (merged cell with namespace name)
            if row[0] and row[0] != "Namespace totals" and row[0] not in ["Totals (main containers)"]:
                # Check if this is a namespace header (merged row) or a workload row
                if row[1] is None and row[2] is None:  # Namespace header
                    current_namespace = row[0]
                    if current_namespace not in namespaces_seen:
                        namespaces_seen.append(current_namespace)
                elif row[1] and current_namespace:  # Workload row - has namespace value in column 2 (index 1)
                    # Verify workload belongs to current namespace
                    assert row[1] == current_namespace, \
                        f"Workload namespace {row[1]} doesn't match current section {current_namespace}"
        
        # Verify we have proper grouping (namespaces should appear only once as section headers)
        assert len(namespaces_seen) == len(set(namespaces_seen)), \
            "Each namespace should appear only once as a section header"


@pytest.fixture
def sample_db(tmp_path):
    """Create a sample database with test data."""
    db_path = tmp_path / 'test.db'
    db = WorkloadDB(str(db_path))
    now = datetime.now(timezone.utc)
    
    # Insert test workloads across multiple namespaces
    manifests = [
        {
            "apiVersion": "apps/v1",
            "kind": "Deployment",
            "metadata": {"name": "app1", "namespace": "ns1"},
            "spec": {
                "replicas": 2,
                "template": {
                    "spec": {
                        "containers": [{
                            "name": "main",
                            "resources": {
                                "requests": {"cpu": "100m", "memory": "128Mi"},
                                "limits": {"cpu": "200m", "memory": "256Mi"}
                            }
                        }]
                    }
                }
            }
        },
        {
            "apiVersion": "apps/v1",
            "kind": "Deployment",
            "metadata": {"name": "app2", "namespace": "ns1"},
            "spec": {
                "replicas": 3,
                "template": {
                    "spec": {
                        "containers": [{
                            "name": "main",
                            "resources": {
                                "requests": {"cpu": "200m", "memory": "256Mi"},
                                "limits": {"cpu": "400m", "memory": "512Mi"}
                            }
                        }]
                    }
                }
            }
        },
        {
            "apiVersion": "apps/v1",
            "kind": "StatefulSet",
            "metadata": {"name": "db1", "namespace": "ns2"},
            "spec": {
                "replicas": 1,
                "template": {
                    "spec": {
                        "containers": [{
                            "name": "postgres",
                            "resources": {
                                "requests": {"cpu": "500m", "memory": "1Gi"},
                                "limits": {"cpu": "1000m", "memory": "2Gi"}
                            }
                        }]
                    }
                }
            }
        }
    ]
    
    for i, manifest in enumerate(manifests):
        db.upsert_workload(
            cluster="test-cluster",
            api_version=manifest["apiVersion"],
            kind=manifest["kind"],
            namespace=manifest["metadata"]["namespace"],
            name=manifest["metadata"]["name"],
            resource_version="1",
            uid=f"u{i+1}",
            manifest=manifest,
            manifest_hash=sha256_of_manifest(manifest),
            now=now
        )
    
    # Insert node capacity data
    db.upsert_node_capacity(
        cluster="test-cluster",
        node_name="worker-1",
        node_data={
            'metadata': {
                'name': 'worker-1',
                'labels': {'node-role.kubernetes.io/worker': ''}
            },
            'status': {
                'capacity': {'cpu': '4000m', 'memory': '8Gi'},
                'allocatable': {'cpu': '3600m', 'memory': '7Gi'},
                'nodeInfo': {
                    'osImage': 'Red Hat Enterprise Linux CoreOS',
                    'kernelVersion': '5.14.0',
                    'containerRuntimeVersion': 'cri-o://1.24.1'
                }
            }
        },
        now=now
    )
    
    return db
