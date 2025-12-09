"""
Tests for Excel formatting in containers config report.

This test file validates that Excel reports apply the same rules as HTML reports,
particularly the request/limit ratio rule (20% threshold).
"""

import pytest
from data_gatherer.reporting.containers_config_report import ContainerConfigurationReport
from data_gatherer.persistence.db import WorkloadDB
from openpyxl import load_workbook


def test_excel_request_limit_ratio_formatting(tmp_path):
    """Test that Excel report applies the 20% request/limit ratio rule."""
    # Create temporary database
    db_path = tmp_path / "test.db"
    db = WorkloadDB(str(db_path))
    
    # Insert test workload with low request/limit ratio (should be flagged)
    low_ratio_manifest = {
        "apiVersion": "apps/v1",
        "kind": "Deployment",
        "metadata": {
            "name": "low-ratio-app",
            "namespace": "test-namespace",
        },
        "spec": {
            "replicas": 1,
            "template": {
                "metadata": {"labels": {"app": "test"}},
                "spec": {
                    "containers": [
                        {
                            "name": "low-ratio-container",
                            "image": "test:latest",
                            "resources": {
                                "requests": {
                                    "cpu": "100m",      # 100m / 1000m = 10% (should be flagged)
                                    "memory": "128Mi"   # 128 / 1024 = 12.5% (should be flagged)
                                },
                                "limits": {
                                    "cpu": "1000m",
                                    "memory": "1024Mi"
                                }
                            },
                            "readinessProbe": {
                                "timeoutSeconds": 5,
                                "initialDelaySeconds": 10
                            }
                        }
                    ]
                }
            }
        }
    }
    
    # Insert test workload with acceptable request/limit ratio (should not be flagged)
    good_ratio_manifest = {
        "apiVersion": "apps/v1",
        "kind": "Deployment",
        "metadata": {
            "name": "good-ratio-app",
            "namespace": "test-namespace",
        },
        "spec": {
            "replicas": 1,
            "template": {
                "metadata": {"labels": {"app": "test"}},
                "spec": {
                    "containers": [
                        {
                            "name": "good-ratio-container",
                            "image": "test:latest",
                            "resources": {
                                "requests": {
                                    "cpu": "500m",     # 500m / 1000m = 50% (should not be flagged)
                                    "memory": "512Mi"  # 512 / 1024 = 50% (should not be flagged)
                                },
                                "limits": {
                                    "cpu": "1000m",
                                    "memory": "1024Mi"
                                }
                            },
                            "readinessProbe": {
                                "timeoutSeconds": 5,
                                "initialDelaySeconds": 10
                            }
                        }
                    ]
                }
            }
        }
    }
    
    # Insert test data
    db.upsert_workload(
        cluster='test-cluster',
        api_version='apps/v1',
        kind='Deployment',
        namespace='test-namespace',
        name='low-ratio-app',
        resource_version='123',
        uid='test-uid-1',
        manifest=low_ratio_manifest,
        manifest_hash='test-hash-1'
    )
    
    db.upsert_workload(
        cluster='test-cluster',
        api_version='apps/v1',
        kind='Deployment',
        namespace='test-namespace',
        name='good-ratio-app',
        resource_version='124',
        uid='test-uid-2',
        manifest=good_ratio_manifest,
        manifest_hash='test-hash-2'
    )
    
    # Generate Excel report
    report = ContainerConfigurationReport()
    output_path = tmp_path / "containers-config-test.xlsx"
    
    report.generate(db, "test-cluster", str(output_path), format='excel')
    
    # Verify report was generated
    assert output_path.exists()
    
    # Load and verify Excel formatting
    wb = load_workbook(str(output_path))
    ws = wb.active
    
    # Find the header row to locate column indices
    headers = []
    for col_idx in range(1, 20):
        cell_value = ws.cell(row=3, column=col_idx).value
        if cell_value:
            headers.append(cell_value)
        else:
            break
    
    # Find column indices for CPU_req_m and Mem_req_Mi
    cpu_req_col = headers.index("CPU_req_m") + 1 if "CPU_req_m" in headers else None
    mem_req_col = headers.index("Mem_req_Mi") + 1 if "Mem_req_Mi" in headers else None
    
    assert cpu_req_col is not None, "CPU_req_m column not found"
    assert mem_req_col is not None, "Mem_req_Mi column not found"
    
    # Find rows for our test containers
    low_ratio_row = None
    good_ratio_row = None
    
    for row_idx in range(4, ws.max_row + 1):
        container_name = ws.cell(row=row_idx, column=4).value  # Container column
        if container_name == "low-ratio-container":
            low_ratio_row = row_idx
        elif container_name == "good-ratio-container":
            good_ratio_row = row_idx
    
    assert low_ratio_row is not None, "low-ratio-container not found in Excel"
    assert good_ratio_row is not None, "good-ratio-container not found in Excel"
    
    # Verify low ratio cells are highlighted (WARNING_MISCONF)
    cpu_req_cell_low = ws.cell(row=low_ratio_row, column=cpu_req_col)
    mem_req_cell_low = ws.cell(row=low_ratio_row, column=mem_req_col)
    
    # The cells should have warning formatting (yellow background or orange text)
    # WARNING_MISCONF uses orange text color (856404)
    assert cpu_req_cell_low.font.color is not None, "CPU request cell should have colored text for low ratio"
    # openpyxl may use different alpha channel prefixes (00, FF, or none)
    assert cpu_req_cell_low.font.color.rgb in ["FF856404", "00856404", "856404"], \
        f"CPU request cell should have warning text color, got {cpu_req_cell_low.font.color.rgb}"
    
    assert mem_req_cell_low.font.color is not None, "Memory request cell should have colored text for low ratio"
    assert mem_req_cell_low.font.color.rgb in ["FF856404", "00856404", "856404"], \
        f"Memory request cell should have warning text color, got {mem_req_cell_low.font.color.rgb}"
    
    # Verify cells have comments explaining the issue
    assert cpu_req_cell_low.comment is not None, "CPU request cell should have a comment"
    assert "20%" in cpu_req_cell_low.comment.text, "Comment should mention 20% threshold"
    
    assert mem_req_cell_low.comment is not None, "Memory request cell should have a comment"
    assert "20%" in mem_req_cell_low.comment.text, "Comment should mention 20% threshold"
    
    # Verify good ratio cells are NOT highlighted
    cpu_req_cell_good = ws.cell(row=good_ratio_row, column=cpu_req_col)
    mem_req_cell_good = ws.cell(row=good_ratio_row, column=mem_req_col)
    
    # These cells should not have warning formatting
    # They might have no font color or default color
    if cpu_req_cell_good.font.color:
        assert cpu_req_cell_good.font.color.rgb not in ["FF856404", "00856404", "856404"], \
            "CPU request cell should not have warning color for good ratio"
    
    if mem_req_cell_good.font.color:
        assert mem_req_cell_good.font.color.rgb not in ["FF856404", "00856404", "856404"], \
            "Memory request cell should not have warning color for good ratio"


def test_excel_missing_values_formatting(tmp_path):
    """Test that Excel report applies formatting for missing values."""
    # Create temporary database
    db_path = tmp_path / "test.db"
    db = WorkloadDB(str(db_path))
    
    # Insert test workload with missing requests and limits
    missing_values_manifest = {
        "apiVersion": "apps/v1",
        "kind": "Deployment",
        "metadata": {
            "name": "missing-values-app",
            "namespace": "test-namespace",
        },
        "spec": {
            "replicas": 1,
            "template": {
                "metadata": {"labels": {"app": "test"}},
                "spec": {
                    "containers": [
                        {
                            "name": "missing-values-container",
                            "image": "test:latest",
                            # No resources defined - should be flagged
                        }
                    ]
                }
            }
        }
    }
    
    db.upsert_workload(
        cluster='test-cluster',
        api_version='apps/v1',
        kind='Deployment',
        namespace='test-namespace',
        name='missing-values-app',
        resource_version='125',
        uid='test-uid-3',
        manifest=missing_values_manifest,
        manifest_hash='test-hash-3'
    )
    
    # Generate Excel report
    report = ContainerConfigurationReport()
    output_path = tmp_path / "containers-config-missing-test.xlsx"
    
    report.generate(db, "test-cluster", str(output_path), format='excel')
    
    # Load and verify Excel formatting
    wb = load_workbook(str(output_path))
    ws = wb.active
    
    # Find the header row
    headers = []
    for col_idx in range(1, 20):
        cell_value = ws.cell(row=3, column=col_idx).value
        if cell_value:
            headers.append(cell_value)
        else:
            break
    
    # Find column indices
    cpu_req_col = headers.index("CPU_req_m") + 1
    cpu_lim_col = headers.index("CPU_lim_m") + 1
    mem_req_col = headers.index("Mem_req_Mi") + 1
    mem_lim_col = headers.index("Mem_lim_Mi") + 1
    
    # Find the row for our test container
    missing_row = None
    for row_idx in range(4, ws.max_row + 1):
        container_name = ws.cell(row=row_idx, column=4).value
        if container_name == "missing-values-container":
            missing_row = row_idx
            break
    
    assert missing_row is not None, "missing-values-container not found in Excel"
    
    # Verify missing request cells have error formatting (red background)
    cpu_req_cell = ws.cell(row=missing_row, column=cpu_req_col)
    mem_req_cell = ws.cell(row=missing_row, column=mem_req_col)
    
    # ERROR_MISS uses red background (F8D7DA)
    # openpyxl may use different alpha channel prefixes (00, FF, or none)
    assert cpu_req_cell.fill.start_color.rgb in ["FFF8D7DA", "00F8D7DA", "F8D7DA"], \
        f"CPU request cell should have error background, got {cpu_req_cell.fill.start_color.rgb}"
    assert mem_req_cell.fill.start_color.rgb in ["FFF8D7DA", "00F8D7DA", "F8D7DA"], \
        f"Memory request cell should have error background, got {mem_req_cell.fill.start_color.rgb}"
    
    # Verify missing limit cells have warning formatting (yellow background)
    cpu_lim_cell = ws.cell(row=missing_row, column=cpu_lim_col)
    mem_lim_cell = ws.cell(row=missing_row, column=mem_lim_col)
    
    # WARNING_MISS uses yellow background (FFF3CD)
    assert cpu_lim_cell.fill.start_color.rgb in ["FFFFF3CD", "00FFF3CD", "FFF3CD"], \
        f"CPU limit cell should have warning background, got {cpu_lim_cell.fill.start_color.rgb}"
    assert mem_lim_cell.fill.start_color.rgb in ["FFFFF3CD", "00FFF3CD", "FFF3CD"], \
        f"Memory limit cell should have warning background, got {mem_lim_cell.fill.start_color.rgb}"
