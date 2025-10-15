from datetime import datetime, timezone
import re
import pytest
from data_gatherer.reporting.cluster_capacity_report import ClusterCapacityReport
from data_gatherer.persistence.db import WorkloadDB
from data_gatherer.util.hash import sha256_of_manifest


def _insert_workload(db: WorkloadDB, cluster: str, kind: str, namespace: str, name: str, replicas: int,
                      cpu_req: str, mem_req: str, cpu_lim: str | None = None, mem_lim: str | None = None):
    now = datetime.now(timezone.utc)
    manifest = {
        'apiVersion': 'apps/v1', 'kind': kind,
        'metadata': {'name': name, 'namespace': namespace},
        'spec': {'replicas': replicas, 'template': {'spec': {'containers': [
            {
                'name': 'c',
                'resources': {
                    'requests': {'cpu': cpu_req, 'memory': mem_req},
                    'limits': {k: v for k, v in [('cpu', cpu_lim), ('memory', mem_lim)] if v}
                }
            }
        ]}}}
    }
    db.upsert_workload(cluster, 'apps/v1', kind, namespace, name, '1', 'u', manifest, sha256_of_manifest(manifest), now)


def test_capacity_report_without_node_data_html(tmp_path):
    """Workloads present but no node_capacity rows -> percentages should be N/A and a notice shown."""
    db_path = tmp_path / 'ns.db'
    db = WorkloadDB(str(db_path))
    # Insert two namespace workloads, but DO NOT insert node capacity rows
    _insert_workload(db, 'c1', 'Deployment', 'team-a', 'api', 2, '200m', '256Mi', '400m', '512Mi')
    _insert_workload(db, 'c1', 'StatefulSet', 'team-b', 'db', 1, '300m', '512Mi', '600m', '1Gi')

    report = ClusterCapacityReport()
    data = report._generate_capacity_data(db, 'c1')
    # Simulate namespace-scoped scenario (no nodes) -> node capacity zero
    assert data['node_capacity']['total_cpu_alloc'] == 0
    assert data['node_capacity']['total_mem_alloc'] == 0
    assert data['node_capacity'].get('namespace_scoped') is True

    html_doc = report._generate_html_report('Cluster Capacity Report: c1', data, 'c1')
    assert 'No worker node capacity data' in html_doc
    # Totals table should contain N/A percentages
    assert re.search(r'Main Containers Requests.*N/A.*N/A', html_doc, re.DOTALL) or 'N/A' in html_doc
    # Ensure namespace rows still rendered
    assert 'team-a' in html_doc and 'team-b' in html_doc


def test_capacity_report_without_node_data_excel(tmp_path):
    pytest.importorskip('openpyxl')
    db_path = tmp_path / 'ns.db'
    db = WorkloadDB(str(db_path))
    _insert_workload(db, 'c1', 'Deployment', 'team-a', 'api', 1, '100m', '128Mi')
    report = ClusterCapacityReport()
    out = tmp_path / 'cap.xlsx'
    report.generate(db, 'c1', str(out), 'excel')
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    ws = wb.active
    # Find the namespace capacity section message when allocatable is zero
    found_notice = any(cell.value == 'No worker node capacity data' for row in ws.iter_rows(min_row=1, max_row=100, max_col=7) for cell in row)
    assert found_notice, 'Expected no-node-capacity notice not found in Excel output'
    # Percent columns should show N/A (search in sheet values)
    # Check that 'N/A' appears in the percentage columns (e.g., columns 6 and 7) of the relevant rows
    na_found = False
    for row in ws.iter_rows(min_row=1, max_row=60, max_col=7):
        # Assuming percentage columns are at index 5 and 6 (6th and 7th columns)
        if any(str(row[i].value) == 'N/A' for i in [5, 6] if len(row) > i):
            na_found = True
            break
    assert na_found, "Expected 'N/A' in percentage columns for capacity"
