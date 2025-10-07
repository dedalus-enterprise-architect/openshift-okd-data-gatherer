import os
import tempfile
import pytest
from data_gatherer.reporting.cluster_capacity_report import ClusterCapacityReport
from data_gatherer.persistence.db import WorkloadDB

def test_cluster_capacity_report_excel(tmp_path):
    # Setup: create a minimal DB with fake data
    db_path = tmp_path / "test.db"
    db = WorkloadDB(str(db_path))
    # Insert minimal node and workload data
    cur = db._conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS node_capacity (
            cluster TEXT, node_name TEXT, cpu_allocatable TEXT, memory_allocatable TEXT, cpu_capacity TEXT, memory_capacity TEXT, node_role TEXT, deleted INTEGER DEFAULT 0, first_seen TEXT, last_seen TEXT
        )
    """)
    cur.execute("""
        INSERT INTO node_capacity (cluster, node_name, cpu_allocatable, memory_allocatable, cpu_capacity, memory_capacity, node_role, deleted, first_seen, last_seen)
        VALUES ('testcluster', 'n1', '2000', '4096', '2000', '4096', 'worker', 0, '2025-10-02T00:00:00Z', '2025-10-02T00:00:00Z')
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS workload (
            cluster TEXT, kind TEXT, namespace TEXT, name TEXT, api_version TEXT, manifest_json TEXT, manifest_hash TEXT, deleted INTEGER DEFAULT 0, first_seen TEXT, last_seen TEXT
        )
    """)
    manifest = '{"kind":"Deployment","spec":{"replicas":2,"template":{"spec":{"containers":[{"name":"c1","resources":{"requests":{"cpu":"500m","memory":"256Mi"},"limits":{"cpu":"1000m","memory":"512Mi"}}}]}}}}'
    cur.execute("""
        INSERT INTO workload (cluster, kind, namespace, name, api_version, manifest_json, manifest_hash, deleted, first_seen, last_seen)
        VALUES ('testcluster', 'Deployment', 'ns1', 'w1', 'apps/v1', ?, 'dummyhash', 0, '2025-10-02T00:00:00Z', '2025-10-02T00:00:00Z')
    """, (manifest,))
    db._conn.commit()

    # Generate Excel report
    out_path = tmp_path / "cluster_capacity_report.xlsx"
    report = ClusterCapacityReport()
    report.generate(db, 'testcluster', str(out_path), format='excel')
    assert os.path.exists(out_path)

    # Validate Excel file contents
    from openpyxl import load_workbook
    wb = load_workbook(str(out_path))
    ws = wb.active
    # Check title
    assert ws.title == "Cluster Capacity Report"
    # Locate summary headers row (contains 'Scope')
    scope_row = None
    for r in range(1, 30):
        if ws.cell(row=r, column=1).value == 'Scope':
            scope_row = r
            break
    assert scope_row, 'Scope header row not found'
    assert [ws.cell(row=scope_row, column=c).value for c in range(1,6)] == [
        'Scope', 'CPU (m)', 'CPU % Allocatable', 'Memory (Mi)', 'Memory % Allocatable'
    ]
    # Locate namespace capacity headers row (first cell 'Namespace')
    ns_header_row = None
    for r in range(scope_row+1, scope_row+80):
        if ws.cell(row=r, column=1).value == 'Namespace':
            ns_header_row = r
            break
    assert ns_header_row, 'Namespace capacity header row not found'
    assert [ws.cell(row=ns_header_row, column=c).value for c in range(1,8)] == [
        'Namespace', 'CPU Requests (m)', 'Memory Requests (Mi)', 'CPU Limits (m)', 'Memory Limits (Mi)', '% CPU allocated on Cluster', '% Memory allocated on Cluster'
    ]
    # Totals row after namespace rows
    totals_found = False
    for r in range(ns_header_row+1, ns_header_row+50):
        if ws.cell(row=r, column=1).value == 'Totals':
            totals_found = True
            break
    assert totals_found, 'Totals row not found in namespace capacity section'
    # Detail section header (Kind / Workload Name)
    detail_found = False
    for r in range(ns_header_row+1, ns_header_row+200):
        if ws.cell(row=r, column=1).value == 'Kind' and ws.cell(row=r, column=2).value == 'Workload Name':
            detail_found = True
            break
    assert detail_found, 'Per-namespace detail header not found'
    wb.close()
