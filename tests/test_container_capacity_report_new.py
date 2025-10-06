"""
Tests for the refactored container capacity report with modular design.

Tests cover:
- Multi-format support (HTML/Excel)
- Data generation and processing
- Resource aggregation
- Edge cases and error handling
"""
import tempfile
import pytest
from pathlib import Path
from datetime import datetime, timezone

from data_gatherer.reporting.container_capacity_report import CapacityReport
from data_gatherer.persistence.db import WorkloadDB
from data_gatherer.util.hash import sha256_of_manifest


@pytest.fixture
def sample_db_with_data(tmp_path):
    """Create a test database with sample workload data."""
    db_path = tmp_path / 'test.db'
    db = WorkloadDB(str(db_path))
    now = datetime.now(timezone.utc)
    
    # Sample Deployment with main and init containers
    deployment_manifest = {
        'apiVersion': 'apps/v1',
        'kind': 'Deployment',
        'metadata': {'name': 'web-app', 'namespace': 'production'},
        'spec': {
            'replicas': 3,
            'template': {
                'spec': {
                    'initContainers': [
                        {
                            'name': 'init-db',
                            'image': 'busybox',
                            'resources': {
                                'requests': {'cpu': '100m', 'memory': '128Mi'},
                                'limits': {'cpu': '200m', 'memory': '256Mi'}
                            }
                        }
                    ],
                    'containers': [
                        {
                            'name': 'web',
                            'image': 'nginx',
                            'resources': {
                                'requests': {'cpu': '200m', 'memory': '256Mi'},
                                'limits': {'cpu': '400m', 'memory': '512Mi'}
                            }
                        },
                        {
                            'name': 'sidecar',
                            'image': 'proxy',
                            'resources': {
                                'requests': {'cpu': '50m', 'memory': '64Mi'}
                                # No limits specified - should trigger conditional formatting
                            }
                        }
                    ]
                }
            }
        }
    }
    
    # Sample StatefulSet in different namespace
    statefulset_manifest = {
        'apiVersion': 'apps/v1',
        'kind': 'StatefulSet',
        'metadata': {'name': 'database', 'namespace': 'storage'},
        'spec': {
            'replicas': 2,
            'template': {
                'spec': {
                    'containers': [
                        {
                            'name': 'db',
                            'image': 'postgres',
                            'resources': {
                                'requests': {'cpu': '500m', 'memory': '1Gi'},
                                'limits': {'cpu': '1000m', 'memory': '2Gi'}
                            }
                        }
                    ]
                }
            }
        }
    }
    
    # Sample Job with missing resources
    job_manifest = {
        'apiVersion': 'batch/v1',
        'kind': 'Job',
        'metadata': {'name': 'batch-process', 'namespace': 'production'},
        'spec': {
            'parallelism': 1,
            'template': {
                'spec': {
                    'containers': [
                        {
                            'name': 'processor',
                            'image': 'batch-app'
                            # No resources specified - should trigger conditional formatting
                        }
                    ]
                }
            }
        }
    }
    
    # Insert workloads
    db.upsert_workload(
        cluster='test-cluster', api_version='apps/v1', kind='Deployment',
        namespace='production', name='web-app', resource_version='1', uid='u1',
        manifest=deployment_manifest, manifest_hash=sha256_of_manifest(deployment_manifest), now=now
    )
    
    db.upsert_workload(
        cluster='test-cluster', api_version='apps/v1', kind='StatefulSet',
        namespace='storage', name='database', resource_version='1', uid='u2',
        manifest=statefulset_manifest, manifest_hash=sha256_of_manifest(statefulset_manifest), now=now
    )
    
    db.upsert_workload(
        cluster='test-cluster', api_version='batch/v1', kind='Job',
        namespace='production', name='batch-process', resource_version='1', uid='u3',
        manifest=job_manifest, manifest_hash=sha256_of_manifest(job_manifest), now=now
    )
    
    return db, str(db_path)


@pytest.fixture
def empty_db(tmp_path):
    """Create an empty test database."""
    db_path = tmp_path / 'empty.db'
    db = WorkloadDB(str(db_path))
    return db, str(db_path)


class TestContainerCapacityReportDataGeneration:
    """Test the core data generation functionality."""
    
    def test_generate_capacity_data_with_data(self, sample_db_with_data):
        """Test data generation with sample workloads."""
        db, _ = sample_db_with_data
        report = CapacityReport()
        
        table_data = report._generate_capacity_data(db, 'test-cluster')
        
        # Verify structure
        assert 'table_rows' in table_data
        assert 'headers' in table_data
        assert 'aggregates' in table_data
        assert 'node_capacity' in table_data
        
        # Verify headers
        expected_headers = [
            "Kind", "Namespace", "Name", "Container", "Type", "Replicas",
            "CPU_req_m", "CPU_lim_m", "Mem_req_Mi", "Mem_lim_Mi",
            "CPU_req_m_total", "CPU_lim_m_total", "Mem_req_Mi_total", "Mem_lim_Mi_total"
        ]
        assert table_data['headers'] == expected_headers
        
        # Verify we have data rows
        table_rows = table_data['table_rows']
        assert len(table_rows) > 0
        
        # Verify row structure - each row should have 14 columns
        for row in table_rows:
            assert len(row) == 14
        
        # Verify aggregates structure
        aggregates = table_data['aggregates']
        expected_agg_keys = [
            'main_cpu_raw', 'main_mem_raw', 'all_cpu_raw', 'all_mem_raw',
            'main_cpu_total', 'main_mem_total', 'all_cpu_total', 'all_mem_total',
            'main_cpu_lim_raw', 'main_mem_lim_raw', 'all_cpu_lim_raw', 'all_mem_lim_raw',
            'main_cpu_lim_total', 'main_mem_lim_total', 'all_cpu_lim_total', 'all_mem_lim_total'
        ]
        for key in expected_agg_keys:
            assert key in aggregates
            assert isinstance(aggregates[key], int)
        
        # Verify node capacity structure
        node_capacity = table_data['node_capacity']
        expected_node_keys = ['worker_cpu_cap', 'worker_cpu_alloc', 'worker_mem_cap', 'worker_mem_alloc']
        for key in expected_node_keys:
            assert key in node_capacity
            assert isinstance(node_capacity[key], int)
    
    def test_generate_capacity_data_empty(self, empty_db):
        """Test data generation with empty database."""
        db, _ = empty_db
        report = CapacityReport()
        
        table_data = report._generate_capacity_data(db, 'empty-cluster')
        
        # Should have structure but empty data
        assert table_data['table_rows'] == []
        assert len(table_data['headers']) == 14
        
        # Aggregates should be all zeros
        aggregates = table_data['aggregates']
        for value in aggregates.values():
            assert value == 0
    
    def test_process_workload_record(self, sample_db_with_data):
        """Test individual workload record processing."""
        db, _ = sample_db_with_data
        report = CapacityReport()

        # Get a sample record
        from data_gatherer.persistence.workload_queries import WorkloadQueries
        from data_gatherer.reporting.common import CONTAINER_WORKLOAD_KINDS

        wq = WorkloadQueries(db)
        rows = wq.list_for_kinds('test-cluster', list(CONTAINER_WORKLOAD_KINDS))

        assert len(rows) > 0
        sample_rec = rows[0]  # Get first workload

        aggregates = report._initialize_aggregates()
        processed_rows = report._process_workload_record(sample_rec, aggregates, worker_node_count=1)

        assert processed_rows is not None
        assert len(processed_rows) > 0

        # Each processed row should have correct structure
        for row in processed_rows:
            assert len(row) == 14
            assert row[0] in ['Deployment', 'StatefulSet', 'Job']  # Kind
            assert row[1] in ['production', 'storage']  # Namespace
            assert isinstance(row[5], int)  # Replicas


class TestContainerCapacityReportHTMLGeneration:
    """Test HTML report generation."""
    
    def test_generate_html_report_with_data(self, sample_db_with_data, tmp_path):
        """Test HTML generation with sample data."""
        db, _ = sample_db_with_data
        report = CapacityReport()
        
        # Generate data and HTML
        table_data = report._generate_capacity_data(db, 'test-cluster')
        html_content = report._generate_html_report('Test Report', table_data, 'test-cluster', db)
        
        # Basic structure checks
        assert '<html>' in html_content
        assert '<h1>Test Report</h1>' in html_content
        assert 'Resource Summary by Namespace' in html_content
        assert 'Resource Summary (Cluster-wide)' in html_content
        
        # Check for table headers
        expected_headers = ["Kind", "Namespace", "Name", "Container", "Type", "Replicas"]
        for header in expected_headers:
            assert f'<th>{header}</th>' in html_content
        
        # Check for namespace grouping
        assert 'production' in html_content
        assert 'storage' in html_content
        
        # Check for totals sections
        assert 'Totals (main containers)' in html_content
        assert 'Totals (all containers incl. init)' in html_content
        assert 'Overhead (init containers)' in html_content
    
    def test_generate_html_report_empty(self, empty_db, tmp_path):
        """Test HTML generation with empty data."""
        db, _ = empty_db
        report = CapacityReport()
        
        table_data = report._generate_capacity_data(db, 'empty-cluster')
        html_content = report._generate_html_report('Empty Report', table_data, 'empty-cluster', db)
        
        assert '<html>' in html_content
        assert '<h1>Empty Report</h1>' in html_content
        assert 'No container workloads found' in html_content
    
    def test_full_html_generation(self, sample_db_with_data, tmp_path):
        """Test complete HTML file generation."""
        db, _ = sample_db_with_data
        report = CapacityReport()
        
        out_path = tmp_path / 'test-report.html'
        report.generate(db, 'test-cluster', str(out_path), 'html')
        
        # Verify file was created
        assert out_path.exists()
        
        # Verify content
        content = out_path.read_text(encoding='utf-8')
        assert 'Capacity aggregation report: test-cluster' in content
        assert 'production' in content
        assert 'storage' in content


class TestContainerCapacityReportExcelGeneration:
    """Test Excel report generation."""
    
    def test_generate_excel_report_with_data(self, sample_db_with_data, tmp_path):
        """Test Excel generation with sample data."""
        pytest.importorskip("openpyxl")
        
        db, _ = sample_db_with_data
        report = CapacityReport()
        
        out_path = tmp_path / 'test-report.xlsx'
        report.generate(db, 'test-cluster', str(out_path), 'excel')
        
        # Verify file was created
        assert out_path.exists()
        
        # Load and verify Excel content
        from openpyxl import load_workbook
        wb = load_workbook(str(out_path))
        ws = wb.active
        
        # Check title
        assert 'Capacity aggregation report: test-cluster' in str(ws['A1'].value)
        
        # Check headers (row 3)
        expected_headers = [
            "Kind", "Namespace", "Name", "Container", "Type", "Replicas",
            "CPU_req_m", "CPU_lim_m", "Mem_req_Mi", "Mem_lim_Mi",
            "CPU_req_m_total", "CPU_lim_m_total", "Mem_req_Mi_total", "Mem_lim_Mi_total"
        ]
        for col_num, header in enumerate(expected_headers, 1):
            assert ws.cell(row=3, column=col_num).value == header
        
        # Verify we have data rows (starting from row 4)
        data_found = False
        for row in range(4, 20):  # Check reasonable range
            if ws.cell(row=row, column=1).value in ['Deployment', 'StatefulSet', 'Job']:
                data_found = True
                break
        assert data_found, "No workload data found in Excel output"
        
        # Check for totals section
        totals_found = False
        for row in range(4, 50):  # Check larger range for totals
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'Totals' in str(cell_value):
                totals_found = True
                break
        assert totals_found, "Totals section not found in Excel output"
    
    def test_generate_excel_report_empty(self, empty_db, tmp_path):
        """Test Excel generation with empty data."""
        pytest.importorskip("openpyxl")
        
        db, _ = empty_db
        report = CapacityReport()
        
        out_path = tmp_path / 'empty-report.xlsx'
        report.generate(db, 'empty-cluster', str(out_path), 'excel')
        
        # Verify file was created
        assert out_path.exists()
        
        # Load and verify Excel content
        from openpyxl import load_workbook
        wb = load_workbook(str(out_path))
        ws = wb.active
        
        # Check title
        assert 'Capacity aggregation report: empty-cluster' in str(ws['A1'].value)
        
        # Should have headers but no data rows (except possibly totals)
        assert ws.cell(row=3, column=1).value == "Kind"
        
        # Row 4 should be empty or contain totals
        row_4_value = ws.cell(row=4, column=1).value
        assert row_4_value is None or 'Totals' in str(row_4_value)
    
    def test_excel_conditional_formatting(self, sample_db_with_data, tmp_path):
        """Test that Excel conditional formatting is applied for missing resources."""
        pytest.importorskip("openpyxl")
        
        db, _ = sample_db_with_data
        report = CapacityReport()
        
        out_path = tmp_path / 'formatting-test.xlsx'
        report.generate(db, 'test-cluster', str(out_path), 'excel')
        
        # Load Excel file
        from openpyxl import load_workbook
        wb = load_workbook(str(out_path))
        ws = wb.active
        
        # Look for cells with missing values and verify formatting
        missing_resource_found = False
        for row in range(4, 20):
            for col in range(7, 11):  # Resource columns
                cell = ws.cell(row=row, column=col)
                if cell.value == '' or cell.value is None:
                    # Should have conditional formatting applied
                    assert cell.fill is not None
                    missing_resource_found = True
        
        # We should find at least one missing resource in our test data
        assert missing_resource_found, "No missing resources found to test conditional formatting"
    
    def test_excel_openpyxl_import_error(self, sample_db_with_data, tmp_path, monkeypatch):
        """Test proper error handling when openpyxl is not available."""
        db, _ = sample_db_with_data
        report = CapacityReport()
        
        # Mock the _generate_excel_report method to simulate ImportError
        def mock_excel_generation(*args, **kwargs):
            raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl")
        
        monkeypatch.setattr(report, '_generate_excel_report', mock_excel_generation)
        
        out_path = tmp_path / 'should-fail.xlsx'
        
        with pytest.raises(ImportError, match="openpyxl is required"):
            report.generate(db, 'test-cluster', str(out_path), 'excel')


class TestContainerCapacityReportResourceCalculations:
    """Test resource calculation accuracy."""
    
    def test_resource_aggregation_accuracy(self, tmp_path):
        """Test that resource calculations are accurate."""
        db_path = tmp_path / 'calc-test.db'
        db = WorkloadDB(str(db_path))
        now = datetime.now(timezone.utc)
        
        # Create a deployment with known resource values for testing
        test_manifest = {
            'apiVersion': 'apps/v1',
            'kind': 'Deployment',
            'metadata': {'name': 'calc-test', 'namespace': 'test'},
            'spec': {
                'replicas': 3,
                'template': {
                    'spec': {
                        'initContainers': [
                            {
                                'name': 'init',
                                'resources': {
                                    'requests': {'cpu': '100m', 'memory': '100Mi'},
                                    'limits': {'cpu': '200m', 'memory': '200Mi'}
                                }
                            }
                        ],
                        'containers': [
                            {
                                'name': 'main',
                                'resources': {
                                    'requests': {'cpu': '200m', 'memory': '256Mi'},
                                    'limits': {'cpu': '400m', 'memory': '512Mi'}
                                }
                            }
                        ]
                    }
                }
            }
        }
        
        db.upsert_workload(
            cluster='calc-test', api_version='apps/v1', kind='Deployment',
            namespace='test', name='calc-test', resource_version='1', uid='u1',
            manifest=test_manifest, manifest_hash=sha256_of_manifest(test_manifest), now=now
        )
        
        report = CapacityReport()
        table_data = report._generate_capacity_data(db, 'calc-test')
        aggregates = table_data['aggregates']
        
        # Verify main container calculations (200m CPU * 3 replicas = 600m)
        assert aggregates['main_cpu_raw'] == 200
        assert aggregates['main_cpu_total'] == 600
        assert aggregates['main_mem_raw'] == 256
        assert aggregates['main_mem_total'] == 768  # 256 * 3
        
        # Verify main container limits
        assert aggregates['main_cpu_lim_raw'] == 400
        assert aggregates['main_cpu_lim_total'] == 1200  # 400 * 3
        assert aggregates['main_mem_lim_raw'] == 512
        assert aggregates['main_mem_lim_total'] == 1536  # 512 * 3
        
        # Verify all containers include init containers
        assert aggregates['all_cpu_raw'] == 300  # 200 + 100
        assert aggregates['all_cpu_total'] == 900  # 600 + 300
        assert aggregates['all_mem_raw'] == 356   # 256 + 100
        assert aggregates['all_mem_total'] == 1068  # 768 + 300
        
        # Verify all container limits
        assert aggregates['all_cpu_lim_raw'] == 600  # 400 + 200
        assert aggregates['all_cpu_lim_total'] == 1800  # 1200 + 600
        assert aggregates['all_mem_lim_raw'] == 712   # 512 + 200
        assert aggregates['all_mem_lim_total'] == 2136  # 1536 + 600
    
    def test_cpu_parsing_edge_cases(self):
        """Test CPU capacity parsing with various formats."""
        report = CapacityReport()
        
        # Test millicores format
        assert report._parse_cpu_capacity('1500m') == 1500
        assert report._parse_cpu_capacity('100m') == 100
        
        # Test cores format (decimal)
        assert report._parse_cpu_capacity('1.5') == 1500
        assert report._parse_cpu_capacity('0.5') == 500
        assert report._parse_cpu_capacity('2') == 2000
        
        # Test edge cases
        assert report._parse_cpu_capacity('') == 0
        assert report._parse_cpu_capacity(None) == 0
        assert report._parse_cpu_capacity('invalid') == 0
        assert report._parse_cpu_capacity('m') == 0  # Invalid format
    
    def test_memory_parsing_edge_cases(self):
        """Test memory capacity parsing with various formats."""
        report = CapacityReport()
        
        # Test various units
        assert report._parse_memory_capacity('1024Ki') == 1  # 1024 Ki = 1 Mi
        assert report._parse_memory_capacity('256Mi') == 256
        assert report._parse_memory_capacity('2Gi') == 2048  # 2 Gi = 2048 Mi
        assert report._parse_memory_capacity('1024') == 1024  # Assume Mi
        
        # Test edge cases
        assert report._parse_memory_capacity('') == 0
        assert report._parse_memory_capacity(None) == 0
        assert report._parse_memory_capacity('invalid') == 0
        assert report._parse_memory_capacity('Mi') == 0  # Invalid format


class TestContainerCapacityReportErrorHandling:
    """Test error handling and edge cases."""
    
    def test_invalid_manifest_handling(self, tmp_path):
        """Test handling of invalid or malformed manifests."""
        db_path = tmp_path / 'invalid-test.db'
        db = WorkloadDB(str(db_path))
        now = datetime.now(timezone.utc)
        
        # Insert workload with invalid manifest (missing pod spec)
        invalid_manifest = {
            'apiVersion': 'apps/v1',
            'kind': 'Deployment',
            'metadata': {'name': 'invalid', 'namespace': 'test'},
            'spec': {
                'replicas': 1
                # Missing template/spec - should be handled gracefully
            }
        }
        
        db.upsert_workload(
            cluster='test', api_version='apps/v1', kind='Deployment',
            namespace='test', name='invalid', resource_version='1', uid='u1',
            manifest=invalid_manifest, manifest_hash=sha256_of_manifest(invalid_manifest), now=now
        )
        
        report = CapacityReport()
        table_data = report._generate_capacity_data(db, 'test')
        
        # Should handle gracefully - no data rows but no crashes
        assert table_data['table_rows'] == []
        assert all(v == 0 for v in table_data['aggregates'].values())
    
    def test_missing_resources_handling(self, tmp_path):
        """Test handling of containers with missing resource specifications."""
        db_path = tmp_path / 'missing-resources.db'
        db = WorkloadDB(str(db_path))
        now = datetime.now(timezone.utc)
        
        # Container with no resources specified
        manifest = {
            'apiVersion': 'apps/v1',
            'kind': 'Deployment',
            'metadata': {'name': 'no-resources', 'namespace': 'test'},
            'spec': {
                'replicas': 2,
                'template': {
                    'spec': {
                        'containers': [
                            {
                                'name': 'no-resources',
                                'image': 'nginx'
                                # No resources block
                            }
                        ]
                    }
                }
            }
        }
        
        db.upsert_workload(
            cluster='test', api_version='apps/v1', kind='Deployment',
            namespace='test', name='no-resources', resource_version='1', uid='u1',
            manifest=manifest, manifest_hash=sha256_of_manifest(manifest), now=now
        )
        
        report = CapacityReport()
        table_data = report._generate_capacity_data(db, 'test')
        
        # Should have one row with empty resource values
        assert len(table_data['table_rows']) == 1
        row = table_data['table_rows'][0]
        
        # Resource columns should be empty strings
        assert row[6] == ''  # CPU_req_m
        assert row[7] == ''  # CPU_lim_m
        assert row[8] == ''  # Mem_req_Mi
        assert row[9] == ''  # Mem_lim_Mi
        assert row[10] == ''  # CPU_req_m_total
        assert row[11] == ''  # CPU_lim_m_total
        assert row[12] == ''  # Mem_req_Mi_total
        assert row[13] == ''  # Mem_lim_Mi_total
        
        # Aggregates should still be zero
        assert all(v == 0 for v in table_data['aggregates'].values())


# Integration test using the CLI
def test_container_capacity_report_cli_integration(tmp_path):
    """Test the full CLI integration for container capacity reports."""
    from click.testing import CliRunner
    from data_gatherer.run import cli
    
    # Create test database with data
    db_path = tmp_path / 'data.db'
    db = WorkloadDB(str(db_path))
    now = datetime.now(timezone.utc)
    
    test_manifest = {
        'apiVersion': 'apps/v1',
        'kind': 'Deployment',
        'metadata': {'name': 'cli-test', 'namespace': 'default'},
        'spec': {
            'replicas': 1,
            'template': {
                'spec': {
                    'containers': [
                        {
                            'name': 'test',
                            'resources': {
                                'requests': {'cpu': '100m', 'memory': '128Mi'}
                            }
                        }
                    ]
                }
            }
        }
    }
    
    db.upsert_workload(
        cluster='cli-test-cluster', api_version='apps/v1', kind='Deployment',
        namespace='default', name='cli-test', resource_version='1', uid='u1',
        manifest=test_manifest, manifest_hash=sha256_of_manifest(test_manifest), now=now
    )
    db._conn.close()
    
    # Create config file
    cfg_content = f'''storage:
  base_dir: {tmp_path.as_posix()}
  write_manifest_files: false
clusters:
  - name: cli-test-cluster
    credentials:
      host: https://dummy
      verify_ssl: false
    include_kinds: [Deployment]
    parallelism: 2
logging:
  level: INFO
  format: text
'''
    cfg_path = tmp_path / 'cfg.yaml'
    cfg_path.write_text(cfg_content)
    
    # Place database in expected location
    cluster_dir = tmp_path / 'cli-test-cluster'
    cluster_dir.mkdir(exist_ok=True)
    import shutil
    shutil.copy(str(db_path), cluster_dir / 'data.db')
    
    # Test HTML generation
    runner = CliRunner()
    result = runner.invoke(cli, [
        '--config', str(cfg_path), 'report',
        '--cluster', 'cli-test-cluster',
        '--type', 'container-capacity'
    ])
    assert result.exit_code == 0, f"CLI failed: {result.output}"
    
    # Test Excel generation if openpyxl is available
    try:
        import openpyxl
        result = runner.invoke(cli, [
            '--config', str(cfg_path), 'report', 
            '--cluster', 'cli-test-cluster',
            '--type', 'container-capacity',
            '--format', 'excel'
        ])
        assert result.exit_code == 0, f"Excel CLI failed: {result.output}"
    except ImportError:
        # Skip Excel test if openpyxl not available
        pass
