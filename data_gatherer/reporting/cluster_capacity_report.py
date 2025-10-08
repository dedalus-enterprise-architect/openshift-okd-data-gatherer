from __future__ import annotations
import html
from typing import List, Dict, Any, Tuple, Optional
from data_gatherer.reporting.base import ReportGenerator, register
from data_gatherer.persistence.db import WorkloadDB
from data_gatherer.persistence.workload_queries import WorkloadQueries
from data_gatherer.reporting.common import (
    CONTAINER_WORKLOAD_KINDS, cpu_to_milli, mem_to_mi,
    extract_pod_spec, calculate_effective_replicas,
    build_legend_html, wrap_html_document,
    format_cell_with_condition
)


@register
class ClusterCapacityReport(ReportGenerator):
    """
    Generate unified cluster capacity reports with comprehensive resource analysis.
    
    Combines container-level details with namespace aggregation and cluster-wide
    capacity metrics. Supports multiple output formats and provides detailed
    resource allocation, utilization, and configuration analysis.
    """
    type_name = 'cluster-capacity'
    filename_prefix = 'cluster-capacity-'
    supported_formats = ['html', 'excel']

    def generate(self, db: WorkloadDB, cluster: str, out_path: str, format: str = 'html') -> None:
        """
        Generate cluster capacity report in specified format.
        
        Args:
            db: Database connection for workload queries
            cluster: Target cluster name
            out_path: Output file path
            format: Output format ('html' or 'excel')
        """
        # Generate format-agnostic data
        capacity_data = self._generate_capacity_data(db, cluster)
        title = f"Cluster Capacity Report: {cluster}"
        
        if format.lower() == 'excel':
            self._generate_excel_report(title, capacity_data, out_path)
        else:
            # Default to HTML
            html_content = self._generate_html_report(title, capacity_data, cluster)
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

    def _generate_capacity_data(self, db: WorkloadDB, cluster: str) -> Dict[str, Any]:
        """
        Generate core capacity data structure used by all output formats.
        
        Args:
            db: Database connection
            cluster: Cluster name
            
        Returns:
            Dictionary containing namespace totals, node capacity, summary data, and detailed per-namespace breakdowns
        """
        # Process workload data and collect details
        wq = WorkloadQueries(db)
        rows = wq.list_for_kinds(cluster, list(CONTAINER_WORKLOAD_KINDS))
        node_capacity = self._get_node_capacity(db, cluster)
        worker_node_count = node_capacity.get('worker_node_count', 0)

        ns_totals: Dict[str, Dict[str, int]] = {}
        ns_details: Dict[str, List[Dict[str, Any]]] = {}

        for rec in rows:
            kind = rec['kind']
            namespace = rec['namespace'] or ''
            name = rec['name']
            manifest = rec['manifest']
            pod_spec = extract_pod_spec(kind, manifest)
            if not pod_spec:
                continue
            replicas = calculate_effective_replicas(kind, manifest, pod_spec, worker_node_count)
            # Do NOT skip workloads with replicas == 0
            if namespace not in ns_totals:
                ns_totals[namespace] = {'cpu': 0, 'mem': 0, 'cpu_lim': 0, 'mem_lim': 0}
            if namespace not in ns_details:
                ns_details[namespace] = []
            # Process containers (main + init)
            for ctype, cdef in [('main', c) for c in pod_spec.get('containers', [])]:
                res = cdef.get('resources', {})
                req = res.get('requests', {}) or {}
                lim = res.get('limits', {}) or {}
                cpu_req = cpu_to_milli(req.get('cpu')) or 0
                mem_req = mem_to_mi(req.get('memory')) or 0
                cpu_lim = cpu_to_milli(lim.get('cpu')) or 0
                mem_lim = mem_to_mi(lim.get('memory')) or 0
                ns_totals[namespace]['cpu'] += cpu_req * replicas
                ns_totals[namespace]['mem'] += mem_req * replicas
                ns_totals[namespace]['cpu_lim'] += cpu_lim * replicas
                ns_totals[namespace]['mem_lim'] += mem_lim * replicas
                ns_details[namespace].append({
                    'kind': kind,
                    'name': name,
                    'container': cdef.get('name', ''),
                    'replicas': replicas,
                    'cpu_req': cpu_req,
                    'mem_req': mem_req,
                    'cpu_lim': cpu_lim,
                    'mem_lim': mem_lim,
                    'cpu_req_total': cpu_req * replicas,
                    'mem_req_total': mem_req * replicas,
                    'cpu_lim_total': cpu_lim * replicas,
                    'mem_lim_total': mem_lim * replicas,
                })
        summary_totals = self._calculate_summary_totals(ns_totals)
        return {
            'ns_totals': ns_totals,
            'node_capacity': node_capacity,
            'summary_totals': summary_totals,
            'ns_details': ns_details
        }

    def _process_namespace_totals(self, db: WorkloadDB, cluster: str) -> Dict[str, Dict[str, int]]:
        """
        Process workload data to generate namespace-level resource totals.
        
        Enhanced to:
        - Properly handle DaemonSets (count per eligible worker node)
        - Filter workloads that won't run on worker nodes
        
        Args:
            db: Database connection
            cluster: Cluster name
            
        Returns:
            Dictionary mapping namespace names to resource totals
        """
        wq = WorkloadQueries(db)
        rows = wq.list_for_kinds(cluster, list(CONTAINER_WORKLOAD_KINDS))
        
        # Get worker node count for DaemonSet calculations
        node_capacity = self._get_node_capacity(db, cluster)
        worker_node_count = node_capacity.get('worker_node_count', 0)
        
        ns_totals: Dict[str, Dict[str, int]] = {}
        
        for rec in rows:
            kind = rec['kind']
            namespace = rec['namespace'] or ''
            name = rec['name']
            manifest = rec['manifest']
            
            pod_spec = extract_pod_spec(kind, manifest)
            if not pod_spec:
                continue
            
            # Use shared calculation logic for consistency across all reports
            replicas = calculate_effective_replicas(kind, manifest, pod_spec, worker_node_count)
            
            # Do NOT skip workloads with replicas == 0
            
            # Initialize namespace totals if not exists
            if namespace not in ns_totals:
                ns_totals[namespace] = {'cpu': 0, 'mem': 0, 'cpu_lim': 0, 'mem_lim': 0}
            
            # Process containers (main + init)
            for cdef in pod_spec.get('containers', []) + pod_spec.get('initContainers', []):
                self._process_container_resources(cdef, replicas, ns_totals[namespace])
        
        return ns_totals
    
    def _process_container_resources(self, cdef: Dict[str, Any], replicas: int, ns_total: Dict[str, int]) -> None:
        """
        Process individual container resource requirements into namespace totals.
        
        Args:
            cdef: Container definition from pod spec
            replicas: Number of replicas for the workload (including zero)
            ns_total: Namespace totals dictionary to update
        """
        res = cdef.get('resources', {})
        req = res.get('requests', {}) or {}
        lim = res.get('limits', {}) or {}
        
        cpu_req = cpu_to_milli(req.get('cpu')) or 0
        mem_req = mem_to_mi(req.get('memory')) or 0
        cpu_lim = cpu_to_milli(lim.get('cpu')) or 0
        mem_lim = mem_to_mi(lim.get('memory')) or 0
        
        # Multiply by actual replica count (including zero)
        if cpu_req:
            ns_total['cpu'] += cpu_req * replicas
        if mem_req:
            ns_total['mem'] += mem_req * replicas
        if cpu_lim:
            ns_total['cpu_lim'] += cpu_lim * replicas
        if mem_lim:
            ns_total['mem_lim'] += mem_lim * replicas

    def _get_node_capacity(self, db: WorkloadDB, cluster: str) -> Dict[str, int]:
        """
        Retrieve and aggregate worker node capacity information.
        
        Enhanced to identify worker nodes using multiple methods:
        - Explicit 'worker' role
        - Nodes without master/infra roles
        
        Uses allocatable values which already account for:
        - System-reserved resources
        - Eviction thresholds
        As per OpenShift formula: Allocatable = Capacity - system-reserved - eviction-thresholds
        
        Args:
            db: Database connection
            cluster: Cluster name
            
        Returns:
            Dictionary with total CPU and memory allocatable/capacity and worker node count
        """
        try:
            cur = db._conn.cursor()
            # Enhanced query: Include nodes marked as worker OR nodes that are not master/infra
            node_rows = cur.execute(
                """SELECT cpu_allocatable, memory_allocatable, cpu_capacity, memory_capacity, node_role
                   FROM node_capacity
                   WHERE cluster=? AND deleted=0 
                   AND (node_role='worker' OR (node_role NOT IN ('master', 'infra')))""",
                (cluster,)
            ).fetchall()
        except Exception:
            node_rows = []

        total_cpu_alloc = total_mem_alloc = 0
        worker_node_count = len(node_rows)
        
        for cpu_alloc, mem_alloc, cpu_cap, mem_cap, node_role in node_rows:
            total_cpu_alloc += (cpu_to_milli(cpu_alloc) or cpu_to_milli(cpu_cap) or 0)
            total_mem_alloc += (mem_to_mi(mem_alloc) or mem_to_mi(mem_cap) or 0)

        return {
            'total_cpu_alloc': total_cpu_alloc,
            'total_mem_alloc': total_mem_alloc,
            'worker_node_count': worker_node_count
        }

    def _calculate_summary_totals(self, ns_totals: Dict[str, Dict[str, int]]) -> Dict[str, int]:
        """
        Calculate cluster-wide summary totals from namespace data.
        
        Args:
            ns_totals: Namespace totals dictionary
            
        Returns:
            Dictionary with cluster-wide totals
        """
        return {
            'total_req_cpu': sum(v['cpu'] for v in ns_totals.values()),
            'total_req_mem': sum(v['mem'] for v in ns_totals.values()),
            'total_lim_cpu': sum(v['cpu_lim'] for v in ns_totals.values()),
            'total_lim_mem': sum(v['mem_lim'] for v in ns_totals.values())
        }

    def _generate_excel_report(self, title: str, capacity_data: Dict[str, Any], out_path: str) -> None:
        """
        Generate Excel report from processed capacity data.
        
        Args:
            title: Report title
            capacity_data: Processed capacity data
            out_path: Output file path
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl") from e

        # Extract data from processed capacity data
        ns_totals = capacity_data['ns_totals']
        node_capacity = capacity_data['node_capacity']
        summary_totals = capacity_data['summary_totals']
        ns_details = capacity_data.get('ns_details', {})
        total_cpu_alloc = node_capacity['total_cpu_alloc']
        total_mem_alloc = node_capacity['total_mem_alloc']

        wb = Workbook()
        ws = wb.active
        ws.title = "Cluster Capacity Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Use a common light grey for totals rows
        totals_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

        current_row = 2
        # Legend (as a note, not collapsible)
        legend_note = (
            "Legend: Namespace: OpenShift projects; CPU/Memory Requests: Sum of all main containers requests x replica number; "
            "CPU/Memory Limits: Sum of all main containers limits x replica number; % CPU/Memory allocated on Cluster: Percentage of Allocatable resources consumed; "
            "Totals: Aggregated namespace requests & limits (percent uses requests); Container Requests vs Allocatable resources on Worker Nodes: Allocatable baseline, requests, free allocatable, limits."
        )
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = legend_note
        cell.font = Font(italic=True, size=10)
        cell.alignment = Alignment(wrap_text=True)
        current_row += 2

        # Section: Container Requests vs Allocatable resources on Worker Nodes
        ws.cell(row=current_row, column=1, value="Container Requests vs Allocatable resources on Worker Nodes").font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1
        summary_headers = [
            "Scope", "CPU (m)", "CPU % Allocatable", "Memory (Mi)", "Memory % Allocatable"
        ]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        total_req_cpu = summary_totals['total_req_cpu']
        total_req_mem = summary_totals['total_req_mem']
        total_lim_cpu = summary_totals['total_lim_cpu']
        total_lim_mem = summary_totals['total_lim_mem']

        def _pct(v: int, d: int) -> str:
            return 'N/A' if d <= 0 else f"{v / d * 100:.1f}%"

        alloc_cpu_pct = '100.0%' if total_cpu_alloc > 0 else 'N/A'
        alloc_mem_pct = '100.0%' if total_mem_alloc > 0 else 'N/A'
        summary_rows = [
            ["Total resources allocatable on Worker nodes", total_cpu_alloc, alloc_cpu_pct, total_mem_alloc, alloc_mem_pct],
            ["Main Containers Requests", total_req_cpu, _pct(total_req_cpu, total_cpu_alloc), total_req_mem, _pct(total_req_mem, total_mem_alloc)],
            ["Free resources (Allocatable - Requests)", max(0, total_cpu_alloc - total_req_cpu), _pct(max(0, total_cpu_alloc - total_req_cpu), total_cpu_alloc), max(0, total_mem_alloc - total_req_mem), _pct(max(0, total_mem_alloc - total_req_mem), total_mem_alloc)],
            ["Main Containers Limits", total_lim_cpu, _pct(total_lim_cpu, total_cpu_alloc), total_lim_mem, _pct(total_lim_mem, total_mem_alloc)]
        ]
        for row in summary_rows:
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                if col == 1:
                    cell.font = Font(bold=True)
            current_row += 1

        current_row += 2
        # Section: Namespace capacity vs Cluster capacity
        ws.cell(row=current_row, column=1, value="Namespace capacity vs Cluster capacity").font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1
        headers = [
            "Namespace", "CPU Requests (m)", "Memory Requests (Mi)",
            "CPU Limits (m)", "Memory Limits (Mi)", "% CPU allocated on Cluster", "% Memory allocated on Cluster"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        if not ns_totals:
            ws.merge_cells(f'A{current_row}:G{current_row}')
            cell = ws.cell(row=current_row, column=1)
            cell.value = "No workloads found"
            cell.font = Font(italic=True)
            current_row += 1
        elif total_cpu_alloc == 0 and total_mem_alloc == 0:
            ws.merge_cells(f'A{current_row}:G{current_row}')
            cell = ws.cell(row=current_row, column=1)
            cell.value = "No worker node capacity data"
            cell.font = Font(italic=True)
            current_row += 1
        else:
            sorted_ns = sorted(ns_totals.items(), key=lambda x: (
                (x[1]['cpu'] / total_cpu_alloc * 100 if total_cpu_alloc else 0) +
                (x[1]['mem'] / total_mem_alloc * 100 if total_mem_alloc else 0)
            ), reverse=True)
            for ns, totals in sorted_ns:
                cpu = totals['cpu']
                mem = totals['mem']
                cpu_lim_ns = totals['cpu_lim']
                mem_lim_ns = totals['mem_lim']
                cpu_pct = f"{cpu / total_cpu_alloc * 100:.1f}%" if total_cpu_alloc else 'N/A'
                mem_pct = f"{mem / total_mem_alloc * 100:.1f}%" if total_mem_alloc else 'N/A'
                row_values = [ns, cpu, mem, cpu_lim_ns, mem_lim_ns, cpu_pct, mem_pct]
                for col, value in enumerate(row_values, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.border = border
                current_row += 1
            # Totals row
            if ns_totals:
                all_req_cpu = sum(v['cpu'] for v in ns_totals.values())
                all_req_mem = sum(v['mem'] for v in ns_totals.values())
                all_lim_cpu = sum(v['cpu_lim'] for v in ns_totals.values())
                all_lim_mem = sum(v['mem_lim'] for v in ns_totals.values())
                cpu_pct_total = f"{all_req_cpu / total_cpu_alloc * 100:.1f}%" if total_cpu_alloc else 'N/A'
                mem_pct_total = f"{all_req_mem / total_mem_alloc * 100:.1f}%" if total_mem_alloc else 'N/A'
                for col, value in enumerate([
                    "Totals", all_req_cpu, all_req_mem, all_lim_cpu, all_lim_mem, cpu_pct_total, mem_pct_total
                ], 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = Font(bold=True)
                    cell.fill = totals_fill
                    cell.border = border
                current_row += 1

        current_row += 2
        # Section: Namespace detailed tables
        for ns, details in ns_details.items():
            ws.cell(row=current_row, column=1, value=f"Namespace: {ns}").font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
            current_row += 1
            detail_headers = [
                "Kind", "Workload Name", "Container", "Replicas",
                "CPU Request (m)", "Memory Request (Mi)", "CPU Limit (m)", "Memory Limit (Mi)",
                "CPU Request × Replicas", "Memory Request × Replicas", "CPU Limit × Replicas", "Memory Limit × Replicas"
            ]
            for col, header in enumerate(detail_headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            current_row += 1
            ns_cpu_req = ns_mem_req = ns_cpu_lim = ns_mem_lim = ns_cpu_req_total = ns_mem_req_total = ns_cpu_lim_total = ns_mem_lim_total = 0
            for row in details:
                ns_cpu_req += row["cpu_req"]
                ns_mem_req += row["mem_req"]
                ns_cpu_lim += row["cpu_lim"]
                ns_mem_lim += row["mem_lim"]
                ns_cpu_req_total += row["cpu_req_total"]
                ns_mem_req_total += row["mem_req_total"]
                ns_cpu_lim_total += row["cpu_lim_total"]
                ns_mem_lim_total += row["mem_lim_total"]
                row_values = [
                    row["kind"], row["name"], row["container"], row["replicas"],
                    row["cpu_req"], row["mem_req"], row["cpu_lim"], row["mem_lim"],
                    row["cpu_req_total"], row["mem_req_total"], row["cpu_lim_total"], row["mem_lim_total"]
                ]
                for col, value in enumerate(row_values, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.border = border
                current_row += 1
            # Totals row for namespace
            for col, value in enumerate([
                "Totals", "", "", "",
                ns_cpu_req, ns_mem_req, ns_cpu_lim, ns_mem_lim,
                ns_cpu_req_total, ns_mem_req_total, ns_cpu_lim_total, ns_mem_lim_total
            ], 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = Font(bold=True)
                cell.fill = totals_fill
                cell.border = border
            current_row += 2

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(out_path)



    def _generate_html_report(self, title: str, capacity_data: Dict[str, Any], cluster: str) -> str:
        """
        Generate HTML report from processed capacity data.
        
        Args:
            title: Report title
            capacity_data: Processed capacity data
            cluster: Cluster name
            
        Returns:
            Complete HTML document as string
        """
        # Extract data from processed capacity data
        ns_totals = capacity_data['ns_totals']
        node_capacity = capacity_data['node_capacity']
        summary_totals = capacity_data['summary_totals']
        ns_details = capacity_data.get('ns_details', {})

        total_cpu_alloc = node_capacity['total_cpu_alloc']
        total_mem_alloc = node_capacity['total_mem_alloc']

        table = [
            '<table border=1 cellpadding=4 cellspacing=0>',
            '<tr><th>Namespace</th><th>CPU Requests (m)</th><th>Memory Requests (Mi)</th>'
            '<th>CPU Limits (m)</th><th>Memory Limits (Mi)</th>'
            '<th>% CPU allocated on Cluster</th><th>% Memory allocated on Cluster</th></tr>'
        ]
        if not ns_totals:
            table.append('<tr><td colspan="6" style="text-align:center; font-style:italic;">No workloads found</td></tr>')
        elif total_cpu_alloc == 0 and total_mem_alloc == 0:
            table.append('<tr><td colspan="6" style="text-align:center; font-style:italic;">No worker node capacity data</td></tr>')
        else:
            for ns, totals in sorted(ns_totals.items(), key=lambda x: (
                (x[1]['cpu'] / total_cpu_alloc * 100 if total_cpu_alloc else 0) +
                (x[1]['mem'] / total_mem_alloc * 100 if total_mem_alloc else 0)
            ), reverse=True):
                cpu = totals['cpu']
                mem = totals['mem']
                cpu_lim_ns = totals['cpu_lim']
                mem_lim_ns = totals['mem_lim']
                cpu_pct_str = f"{cpu / total_cpu_alloc * 100:.1f}%" if total_cpu_alloc else 'N/A'
                mem_pct_str = f"{mem / total_mem_alloc * 100:.1f}%" if total_mem_alloc else 'N/A'
                table.append(
                    f'<tr><td>{html.escape(ns)}</td><td>{cpu:,}</td><td>{mem:,}</td>'
                    f'<td>{cpu_lim_ns:,}</td><td>{mem_lim_ns:,}</td>'
                    f'<td>{cpu_pct_str}</td><td>{mem_pct_str}</td></tr>'
                )
            # Totals row summarizing all namespaces
            if ns_totals:
                all_req_cpu = summary_totals['total_req_cpu']
                all_req_mem = summary_totals['total_req_mem']
                all_lim_cpu = summary_totals['total_lim_cpu']
                all_lim_mem = summary_totals['total_lim_mem']
                cpu_pct_total = f"{all_req_cpu / total_cpu_alloc * 100:.1f}%" if total_cpu_alloc else 'N/A'
                mem_pct_total = f"{all_req_mem / total_mem_alloc * 100:.1f}%" if total_mem_alloc else 'N/A'
                table.append(
                    '<tr style="background-color: #eaeaea;">'
                    f'<td><strong>Totals</strong></td><td><strong>{all_req_cpu:,}</strong></td>'
                    f'<td><strong>{all_req_mem:,}</strong></td>'
                    f'<td><strong>{all_lim_cpu:,}</strong></td><td><strong>{all_lim_mem:,}</strong></td>'
                    f'<td><strong>{cpu_pct_total}</strong></td><td><strong>{mem_pct_total}</strong></td>'
                    '</tr>'
                )
        table.append('</table>')

        # Namespace detailed tables (new)
        details_tables = []
        for ns, details in ns_details.items():
            details_tables.append(f'<h3>Namespace: {html.escape(ns)}</h3>')
            details_tables.append('<table border=1 cellpadding=4 cellspacing=0>')
            details_tables.append('<tr>'
                '<th>Kind</th><th>Workload Name</th><th>Container</th><th>Replicas</th>'
                '<th>CPU Request (m)</th><th>Memory Request (Mi)</th>'
                '<th>CPU Limit (m)</th><th>Memory Limit (Mi)</th>'
                '<th>CPU Request × Replicas</th><th>Memory Request × Replicas</th>'
                '<th>CPU Limit × Replicas</th><th>Memory Limit × Replicas</th>'
            '</tr>')
            # Accumulate totals for this namespace
            ns_cpu_req = ns_mem_req = ns_cpu_lim = ns_mem_lim = ns_cpu_req_total = ns_mem_req_total = ns_cpu_lim_total = ns_mem_lim_total = 0
            for row in details:
                ns_cpu_req += row["cpu_req"]
                ns_mem_req += row["mem_req"]
                ns_cpu_lim += row["cpu_lim"]
                ns_mem_lim += row["mem_lim"]
                ns_cpu_req_total += row["cpu_req_total"]
                ns_mem_req_total += row["mem_req_total"]
                ns_cpu_lim_total += row["cpu_lim_total"]
                ns_mem_lim_total += row["mem_lim_total"]
                details_tables.append(
                    f'<tr>'
                    f'<td>{html.escape(row["kind"])}</td>'
                    f'<td>{html.escape(row["name"])}</td>'
                    f'<td>{html.escape(row["container"])}</td>'
                    f'<td>{row["replicas"]}</td>'
                    f'<td>{row["cpu_req"]}</td>'
                    f'<td>{row["mem_req"]}</td>'
                    f'<td>{row["cpu_lim"]}</td>'
                    f'<td>{row["mem_lim"]}</td>'
                    f'<td>{row["cpu_req_total"]}</td>'
                    f'<td>{row["mem_req_total"]}</td>'
                    f'<td>{row["cpu_lim_total"]}</td>'
                    f'<td>{row["mem_lim_total"]}</td>'
                    '</tr>'
                )
            # Totals row with balloon tooltips
            details_tables.append(
                '<tr style="background-color: #eaeaea;">'
                '<td colspan="4"><strong>Totals</strong></td>'
                f'<td data-tooltip="Sum of CPU requests for all containers">{ns_cpu_req}</td>'
                f'<td data-tooltip="Sum of Memory requests for all containers">{ns_mem_req}</td>'
                f'<td data-tooltip="Sum of CPU limits for all containers">{ns_cpu_lim}</td>'
                f'<td data-tooltip="Sum of Memory limits for all containers">{ns_mem_lim}</td>'
                f'<td data-tooltip="Sum of CPU requests × replicas">{ns_cpu_req_total}</td>'
                f'<td data-tooltip="Sum of Memory requests × replicas">{ns_mem_req_total}</td>'
                f'<td data-tooltip="Sum of CPU limits × replicas">{ns_cpu_lim_total}</td>'
                f'<td data-tooltip="Sum of Memory limits × replicas">{ns_mem_lim_total}</td>'
                '</tr>'
            )
            details_tables.append('</table>')

        total_req_cpu = summary_totals['total_req_cpu']
        total_req_mem = summary_totals['total_req_mem']
        total_lim_cpu = summary_totals['total_lim_cpu']
        total_lim_mem = summary_totals['total_lim_mem']

        def _pct(v: int, d: int) -> str:
            return 'N/A' if d <= 0 else f"{v / d * 100:.1f}%"

        totals_table: List[str] = [
            '<table border=1 cellpadding=4 cellspacing=0>',
            '<tr><th>Scope</th><th>CPU (m)</th><th>CPU % Allocatable</th><th>Memory (Mi)</th><th>Memory % Allocatable</th></tr>'
        ]
        # Baseline cluster worker allocatable (always 100% when >0)
        alloc_cpu_pct = '100.0%' if total_cpu_alloc > 0 else 'N/A'
        alloc_mem_pct = '100.0%' if total_mem_alloc > 0 else 'N/A'
        totals_table.append(
            '<tr>'
            f'<td>Total resources allocatable on Worker nodes</td><td>{total_cpu_alloc:,}</td><td>{alloc_cpu_pct}</td>'
            f'<td>{total_mem_alloc:,}</td><td>{alloc_mem_pct}</td></tr>'
        )
        totals_table.append(
            '<tr>'
            f'<td>Main Containers Requests</td><td>{total_req_cpu:,}</td><td>{_pct(total_req_cpu, total_cpu_alloc)}</td>'
            f'<td>{total_req_mem:,}</td><td>{_pct(total_req_mem, total_mem_alloc)}</td></tr>'
        )
        # Free allocatable (clamped to 0 for overcommit)
        free_cpu = max(0, total_cpu_alloc - total_req_cpu)
        free_mem = max(0, total_mem_alloc - total_req_mem)
        totals_table.append(
            '<tr>'
            f'<td>Free resources (Allocatable - Requests)</td><td>{free_cpu:,}</td><td>{_pct(free_cpu, total_cpu_alloc)}</td>'
            f'<td>{free_mem:,}</td><td>{_pct(free_mem, total_mem_alloc)}</td></tr>'
        )
        totals_table.append(
            '<tr>'
            f'<td>Main Containers Limits</td><td>{total_lim_cpu:,}</td><td>{_pct(total_lim_cpu, total_cpu_alloc)}</td>'
            f'<td>{total_lim_mem:,}</td><td>{_pct(total_lim_mem, total_mem_alloc)}</td></tr>'
        )
        totals_table.append('</table>')

        legend_sections = [
            {
                'title': 'Columns',
                'items': [
                    "Namespace: OpenShift projects",
                    'CPU/Memory Requests: Sum of all main containers requests x replica number',
                    'CPU/Memory Limits: Sum of all main containers limits x replica number',
                    '% CPU/Memory allocated on Cluster: Percentage of Allocatable resources consumed',
                    'Totals: Aggregated namespace requests & limits (percent uses requests)',
                    'Container Requests vs Allocatable resources on Worker Nodes: Allocatable baseline, requests, free allocatable, limits'
                ]
            }
        ]
        legend_html = build_legend_html(legend_sections)
        # Insert details_tables after the main namespace table and before the totals table
        return wrap_html_document(
            title,
            [
                legend_html,
                '<h2>Container Requests vs Allocatable resources on Worker Nodes</h2>',
                *totals_table,
                '<h2>Namespace capacity vs Cluster capacity</h2>',
                *table,
                *details_tables
            ]
        )

