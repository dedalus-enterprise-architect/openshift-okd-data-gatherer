from __future__ import annotations
import html
from typing import List, Dict, Any, Optional, Tuple
from data_gatherer.reporting.base import ReportGenerator, register
from data_gatherer.persistence.db import WorkloadDB
from data_gatherer.persistence.workload_queries import WorkloadQueries
from data_gatherer.reporting.common import (
    CONTAINER_WORKLOAD_KINDS, cpu_to_milli, mem_to_mi,
    extract_pod_spec, calculate_effective_replicas,
    build_legend_html, wrap_html_document,
    format_cell_with_condition
)


@register
class CapacityReport(ReportGenerator):
    """
    Generate container capacity reports with resource usage analysis.
    
    Supports multiple output formats and provides detailed resource
    aggregation by namespace with cluster-wide summaries.
    """
    # Renamed from 'capacity' to 'container-capacity' (no backward alias requested)
    type_name = 'container-capacity'
    filename_prefix = 'container-capacity-'
    supported_formats = ['html', 'excel']

    def generate(self, db: WorkloadDB, cluster: str, out_path: str, format: str = 'html') -> None:
        """
        Generate capacity report in specified format.
        
        Args:
            db: Database connection for workload queries
            cluster: Target cluster name
            out_path: Output file path
            format: Output format ('html' or 'excel')
        """
        # Generate format-agnostic data
        table_data = self._generate_capacity_data(db, cluster)
        title = f"Capacity aggregation report: {cluster}"
        
        if format.lower() == 'excel':
            self._generate_excel_report(title, table_data, out_path)
        else:
            # Default to HTML
            html_content = self._generate_html_report(title, table_data, cluster, db)
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

    def _generate_capacity_data(self, db: WorkloadDB, cluster: str) -> Dict[str, Any]:
        """
        Generate core capacity data structure used by all output formats.
        
        Args:
            db: Database connection
            cluster: Cluster name
            
        Returns:
            Dictionary containing table rows, headers, and aggregated totals
        """
        wq = WorkloadQueries(db)
        rows = wq.list_for_kinds(cluster, list(CONTAINER_WORKLOAD_KINDS))
        
        # Get worker node count for DaemonSet calculations
        node_capacity = self._get_node_capacity(db, cluster)
        worker_node_count = node_capacity.get('worker_node_count', 0)
        
        table_rows: List[List] = []
        aggregates = self._initialize_aggregates()
        
        for rec in rows:
            processed_rows = self._process_workload_record(rec, aggregates, worker_node_count)
            if processed_rows:
                table_rows.extend(processed_rows)
        
        headers = [
            "Kind", "Namespace", "Name", "Container", "Type", "Replicas",
            "CPU_req_m", "CPU_lim_m", "Mem_req_Mi", "Mem_lim_Mi",
            "CPU_req_m_total", "CPU_lim_m_total", "Mem_req_Mi_total", "Mem_lim_Mi_total"
        ]
        
        return {
            'table_rows': table_rows,
            'headers': headers,
            'aggregates': aggregates,
            'node_capacity': node_capacity
        }

    def _initialize_aggregates(self) -> Dict[str, int]:
        """Initialize aggregate counters for resource tracking.
        Init containers are discarded so we only keep 'main' aggregates.
        """
        return {
            'main_cpu_raw': 0,
            'main_mem_raw': 0,
            'main_cpu_total': 0,
            'main_mem_total': 0,
            'main_cpu_lim_raw': 0,
            'main_mem_lim_raw': 0,
            'main_cpu_lim_total': 0,
            'main_mem_lim_total': 0,
        }

    def _process_workload_record(self, rec: Dict[str, Any], aggregates: Dict[str, int], worker_node_count: int) -> Optional[List[List]]:
        """
        Process a single workload record into table rows.
        
        Args:
            rec: Workload record from database
            aggregates: Aggregate counters to update
            worker_node_count: Number of worker nodes for DaemonSet calculations
            
        Returns:
            List of table rows for this workload, or None if invalid
        """
        kind = rec['kind']
        namespace = rec['namespace']
        name = rec['name']
        manifest = rec['manifest']
        
        pod_spec = extract_pod_spec(kind, manifest)
        if not pod_spec:
            return None

        # Use shared calculation logic for consistency across all reports
        replicas = calculate_effective_replicas(kind, manifest, pod_spec, worker_node_count)

        # Discard init containers entirely (business decision: runtime sizing focuses only on main containers)
        containers = [(c, 'main') for c in pod_spec.get('containers', [])]

        rows = []
        for cdef, ctype in containers:
            row = self._process_container(cdef, ctype, kind, namespace, name, replicas, aggregates)
            rows.append(row)

        return rows

    def _process_container(
        self, 
        cdef: Dict[str, Any], 
        ctype: str, 
        kind: str, 
        namespace: str, 
        name: str, 
        replicas: int,
        aggregates: Dict[str, int]
    ) -> List:
        """
        Process individual container resource requirements.
        
        Args:
            cdef: Container definition from pod spec
            ctype: Container type ('main' or 'init')
            kind: Workload kind
            namespace: Workload namespace
            name: Workload name
            replicas: Number of replicas
            aggregates: Resource aggregates to update
            
        Returns:
            Table row data for this container
        """
        resources = cdef.get('resources', {})
        req = resources.get('requests', {}) or {}
        lim = resources.get('limits', {}) or {}
        
        # Parse resource values
        cpu_req = cpu_to_milli(req.get('cpu'))
        mem_req = mem_to_mi(req.get('memory'))
        cpu_lim = cpu_to_milli(lim.get('cpu'))
        mem_lim = mem_to_mi(lim.get('memory'))
        
        # Calculate totals
        cpu_req_total = cpu_req * replicas if cpu_req is not None else ''
        mem_req_total = mem_req * replicas if mem_req is not None else ''
        cpu_lim_total = cpu_lim * replicas if cpu_lim is not None else ''
        mem_lim_total = mem_lim * replicas if mem_lim is not None else ''
        
        # Update aggregates
        self._update_aggregates(aggregates, ctype, cpu_req, mem_req, cpu_lim, mem_lim,
                               cpu_req_total, mem_req_total, cpu_lim_total, mem_lim_total)
        
        return [
            kind, namespace or '', name, cdef.get('name') or '', ctype,
            replicas,
            '' if cpu_req is None else cpu_req,
            '' if cpu_lim is None else cpu_lim,
            '' if mem_req is None else mem_req,
            '' if mem_lim is None else mem_lim,
            '' if isinstance(cpu_req_total, str) else cpu_req_total,
            '' if isinstance(cpu_lim_total, str) else cpu_lim_total,
            '' if isinstance(mem_req_total, str) else mem_req_total,
            '' if isinstance(mem_lim_total, str) else mem_lim_total,
        ]

    def _update_aggregates(
        self, 
        aggregates: Dict[str, int], 
        ctype: str,
        cpu_req: Optional[int], 
        mem_req: Optional[int],
        cpu_lim: Optional[int], 
        mem_lim: Optional[int],
        cpu_req_total: Any, 
        mem_req_total: Any,
        cpu_lim_total: Any, 
        mem_lim_total: Any
    ) -> None:
        """Update resource aggregates with container values."""
        if ctype != 'main':
            return  # Should not happen because we don't process init containers

        if cpu_req is not None:
            aggregates['main_cpu_raw'] += cpu_req
            if isinstance(cpu_req_total, int):
                aggregates['main_cpu_total'] += cpu_req_total

        if cpu_lim is not None:
            aggregates['main_cpu_lim_raw'] += cpu_lim
            if isinstance(cpu_lim_total, int):
                aggregates['main_cpu_lim_total'] += cpu_lim_total

        if mem_req is not None:
            aggregates['main_mem_raw'] += mem_req
            if isinstance(mem_req_total, int):
                aggregates['main_mem_total'] += mem_req_total

        if mem_lim is not None:
            aggregates['main_mem_lim_raw'] += mem_lim
            if isinstance(mem_lim_total, int):
                aggregates['main_mem_lim_total'] += mem_lim_total

    def _get_node_capacity(self, db: WorkloadDB, cluster: str) -> Dict[str, int]:
        """
        Retrieve and parse worker node capacity information.
        
        Enhanced to identify worker nodes using multiple methods:
        - Explicit 'worker' role
        - Nodes without master/infra roles
        
        Args:
            db: Database connection
            cluster: Cluster name
            
        Returns:
            Dictionary with parsed capacity values and worker node count
        """
        try:
            cur = db._conn.cursor()
            node_rows = cur.execute(
                "SELECT cpu_capacity, cpu_allocatable, memory_capacity, memory_allocatable, node_role "
                "FROM node_capacity WHERE cluster=? AND deleted=0",
                (cluster,)
            ).fetchall()
        except Exception:
            node_rows = []

        worker_cpu_cap = worker_cpu_alloc = worker_mem_cap = worker_mem_alloc = 0
        worker_node_count = 0
        for nr in node_rows:
            node_role = (nr[4] or '').lower()
            # Include worker nodes OR nodes that are not master/infra
            if node_role == 'worker' or (node_role not in ('master', 'infra')):
                worker_cpu_cap += self._parse_cpu_capacity(nr[0])
                worker_cpu_alloc += self._parse_cpu_capacity(nr[1])
                worker_mem_cap += self._parse_memory_capacity(nr[2])
                worker_mem_alloc += self._parse_memory_capacity(nr[3])
                worker_node_count += 1

        return {
            'worker_cpu_cap': worker_cpu_cap,
            'worker_cpu_alloc': worker_cpu_alloc,
            'worker_mem_cap': worker_mem_cap,
            'worker_mem_alloc': worker_mem_alloc,
            'worker_node_count': worker_node_count
        }

    def _parse_cpu_capacity(self, val: Optional[str]) -> int:
        """
        Convert CPU capacity string to millicores.
        
        Args:
            val: CPU value string (e.g., '2', '1500m', '0.5')
            
        Returns:
            CPU value in millicores
        """
        if not val:
            return 0
        v = val.strip()
        if v.endswith('m'):
            try:
                return int(v[:-1])
            except ValueError:
                return 0
        # Assume cores (can be fractional)
        try:
            return int(float(v) * 1000)
        except ValueError:
            return 0

    def _parse_memory_capacity(self, val: Optional[str]) -> int:
        """
        Convert memory capacity string to MiB.
        
        Args:
            val: Memory value string (e.g., '8Gi', '2048Mi', '1024Ki')
            
        Returns:
            Memory value in MiB
        """
        if not val:
            return 0
        v = val.strip()
        try:
            if v.endswith('Ki'):
                return int(v[:-2]) // 1024
            if v.endswith('Mi'):
                return int(v[:-2])
            if v.endswith('Gi'):
                return int(v[:-2]) * 1024
            return int(v)  # assume already MiB
        except ValueError:
            return 0

    def _generate_html_report(self, title: str, table_data: Dict[str, Any], cluster: str, db: WorkloadDB) -> str:
        """Generate HTML report."""
        table_rows = table_data['table_rows']
        headers = table_data['headers']
        aggregates = table_data['aggregates']
        node_capacity = table_data['node_capacity']

        parts = [f"<h1>{html.escape(title)}</h1>"]

        main_cpu_raw = aggregates['main_cpu_raw']
        main_mem_raw = aggregates['main_mem_raw']
        main_cpu_total = aggregates['main_cpu_total']
        main_mem_total = aggregates['main_mem_total']
        main_cpu_lim_raw = aggregates['main_cpu_lim_raw']
        main_mem_lim_raw = aggregates['main_mem_lim_raw']
        main_cpu_lim_total = aggregates['main_cpu_lim_total']
        main_mem_lim_total = aggregates['main_mem_lim_total']

        capacity_legend_sections = [
            {
                'title': 'Columns',
                'items': [
                    'Kind: Workload controller kind (Deployment / StatefulSet / etc.)',
                    'Namespace: Kubernetes namespace (blank for cluster‑scoped)',
                    'Name: Workload name',
                    'Container: Container name inside pod spec',
                    'Type: Always "main"',
                    'Replicas: Replica number set in spec (DaemonSet replicas only on worker nodes)',
                    'CPU_req_m: CPU request (millicores)',
                    'CPU_lim_m: CPU limit (millicores)',
                    'Mem_req_Mi: Memory request (MiB)',
                    'Mem_lim_Mi: Memory limit (MiB)',
                    'CPU_req_m_total: CPU request * replicas',
                    'CPU_lim_m_total: CPU limit * replicas',
                    'Mem_req_Mi_total: Memory request * replicas',
                    'Mem_lim_Mi_total: Memory limit * replicas'
                ]
            },
            {
                'title': 'Configuration Severities',
                'items': [
                    {'class': 'error-miss-cell', 'description': 'Required value or parameter is missing'},
                    {'class': 'warning-miss-cell', 'description': 'Recommended value or parameter is missing'},
                    {'class': 'error-misconf-cell', 'description': 'Wrong value or parameter set'},
                    {'class': 'warning-misconf-cell', 'description': 'Value or parameter set should be re-evaluated'}
                ]
            },
            {
                'title': 'Structural Rows',
                'items': [
                    {'class': 'totals-row-main', 'description': 'Cluster totals (main containers)'},
                    {'class': 'ns-totals', 'description': 'Per-namespace aggregated totals'}
                ]
            }
        ]
        parts.append(build_legend_html(capacity_legend_sections))
        parts.append('<h2>Resource Summary by Namespace</h2>')
        parts.append('<table border=1 cellpadding=4 cellspacing=0>')
        parts.append('<tr>' + ''.join(f'<th>{h}</th>' for h in headers) + '</tr>')

        from itertools import groupby
        table_rows_sorted = sorted(table_rows, key=lambda r: (r[1], r[0], r[2], r[3]))
        for ns, rows_group in groupby(table_rows_sorted, key=lambda r: r[1]):
            rows_list = list(rows_group)
            parts.append(f'<tr><td><strong>{html.escape(ns or "")}</strong></td><td colspan="13"></td></tr>')
            for row in rows_list:
                values = row[:-1] if len(row) > len(headers) else row
                row_cells = []
                for i, col in enumerate(values):
                    if i < len(headers):
                        row_data = {headers[j]: values[j] for j in range(min(len(headers), len(values)))}
                        row_cells.append(format_cell_with_condition(str(col), headers[i], row_data, 'container-capacity'))
                    else:
                        row_cells.append(f'<td>{html.escape(str(col))}</td>')
                parts.append('<tr>' + ''.join(row_cells) + '</tr>')
            if rows_list:
                ns_main_cpu_raw = sum(int(r[6]) for r in rows_list if r[4] == 'main' and r[6] != '')
                ns_main_cpu_lim_raw = sum(int(r[7]) for r in rows_list if r[4] == 'main' and r[7] != '')
                ns_main_mem_raw = sum(int(r[8]) for r in rows_list if r[4] == 'main' and r[8] != '')
                ns_main_mem_lim_raw = sum(int(r[9]) for r in rows_list if r[4] == 'main' and r[9] != '')
                ns_main_cpu_total = sum(int(r[10]) for r in rows_list if r[4] == 'main' and r[10] != '')
                ns_main_cpu_lim_total = sum(int(r[11]) for r in rows_list if r[4] == 'main' and r[11] != '')
                ns_main_mem_total = sum(int(r[12]) for r in rows_list if r[4] == 'main' and r[12] != '')
                ns_main_mem_lim_total = sum(int(r[13]) for r in rows_list if r[4] == 'main' and r[13] != '')
                comments = self._namespace_totals_comments(ns or '', {
                    'CPU_req_m': ns_main_cpu_raw,
                    'CPU_lim_m': ns_main_cpu_lim_raw,
                    'Mem_req_Mi': ns_main_mem_raw,
                    'Mem_lim_Mi': ns_main_mem_lim_raw,
                    'CPU_req_m_total': ns_main_cpu_total,
                    'CPU_lim_m_total': ns_main_cpu_lim_total,
                    'Mem_req_Mi_total': ns_main_mem_total,
                    'Mem_lim_Mi_total': ns_main_mem_lim_total,
                })
                ns_cells = [
                    f'<th colspan="6" title="{html.escape(comments["label"]) }">Namespace totals</th>',
                    f'<th title="{html.escape(comments["CPU_req_m"]) }">{ns_main_cpu_raw}</th>',
                    f'<th title="{html.escape(comments["CPU_lim_m"]) }">{ns_main_cpu_lim_raw}</th>',
                    f'<th title="{html.escape(comments["Mem_req_Mi"]) }">{ns_main_mem_raw}</th>',
                    f'<th title="{html.escape(comments["Mem_lim_Mi"]) }">{ns_main_mem_lim_raw}</th>',
                    f'<th title="{html.escape(comments["CPU_req_m_total"]) }">{ns_main_cpu_total}</th>',
                    f'<th title="{html.escape(comments["CPU_lim_m_total"]) }">{ns_main_cpu_lim_total}</th>',
                    f'<th title="{html.escape(comments["Mem_req_Mi_total"]) }">{ns_main_mem_total}</th>',
                    f'<th title="{html.escape(comments["Mem_lim_Mi_total"]) }">{ns_main_mem_lim_total}</th>'
                ]
                parts.append('<tr>' + ''.join(ns_cells) + '</tr>')

        if table_rows:
            parts.append('<tr class="section-separator"><td colspan="14"></td></tr>')
            totals_cells = [
                '<th colspan="6">Totals (main containers)</th>',
                format_cell_with_condition(str(main_cpu_raw), 'CPU_req_m', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>'),
                format_cell_with_condition(str(main_cpu_lim_raw), 'CPU_lim_m', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>'),
                format_cell_with_condition(str(main_mem_raw), 'Mem_req_Mi', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>'),
                format_cell_with_condition(str(main_mem_lim_raw), 'Mem_lim_Mi', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>'),
                format_cell_with_condition(str(main_cpu_total), 'CPU_req_m_total', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>'),
                format_cell_with_condition(str(main_cpu_lim_total), 'CPU_lim_m_total', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>'),
                format_cell_with_condition(str(main_mem_total), 'Mem_req_Mi_total', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>'),
                format_cell_with_condition(f'{main_mem_lim_total} ({main_mem_lim_total/1024:.2f} GiB)', 'Mem_lim_Mi_total', None, 'container-capacity').replace('<td', '<th').replace('</td>', '</th>')
            ]
            parts.append('<tr>' + ''.join(totals_cells) + '</tr>')
        parts.append('</table>')
        if not table_rows:
            parts.append('<p>No container workloads found for this cluster.</p>')

        worker_cpu_cap = node_capacity['worker_cpu_cap']
        worker_cpu_alloc = node_capacity['worker_cpu_alloc']
        worker_mem_cap = node_capacity['worker_mem_cap']
        worker_mem_alloc = node_capacity['worker_mem_alloc']

        parts.append('<h2>Resource Summary (Cluster-wide)</h2>')
        parts.append('<h3>Containers</h3>')
        parts.append('<table border=1 cellpadding=4 cellspacing=0>')
        parts.append('<tr><th>Scope</th><th>CPU Requests (m)</th><th>CPU Limits (m)</th><th>Memory Requests (Mi)</th><th>Memory Limits (Mi)</th></tr>')
        parts.append('<tr>'
                     f'<td>Containers</td>'
                     f'<td>{main_cpu_total}</td>'
                     f'<td>{main_cpu_lim_total}</td>'
                     f'<td>{main_mem_total}</td>'
                     f'<td>{main_mem_lim_total}</td>'
                     '</tr>')
        parts.append('</table>')

        parts.append('<h3>Worker Nodes</h3>')
        parts.append('<table border=1 cellpadding=4 cellspacing=0>')
        parts.append('<tr><th>CPU Capacity (m)</th><th>CPU Allocatable (m)</th><th>Memory Capacity (Mi)</th><th>Memory Allocatable (Mi)</th></tr>')
        parts.append('<tr>'
                     f'<td>{worker_cpu_cap}</td>'
                     f'<td>{worker_cpu_alloc}</td>'
                     f'<td>{worker_mem_cap}</td>'
                     f'<td>{worker_mem_alloc}</td>'
                     '</tr>')
        parts.append('</table>')

        additional_css = (
            " .totals-row-main th { background:#dfe; }"
            " .ns-totals th { background:#f5f5f5; }"
            " .legend-box.totals-row-main { background: #dfe; }"
            " .legend-box.ns-totals { background: #f5f5f5; }"
            " .section-separator td { background:#fff; border:0; height:14px; }"
        )
        return wrap_html_document(title, parts, additional_css)

    def _generate_excel_report(self, title: str, table_data: Dict[str, Any], out_path: str) -> None:
        """
        Generate Excel report with conditional formatting and professional styling.
        
        Args:
            title: Report title
            table_data: Processed table data
            out_path: Output file path
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Container Capacity"
        
        # Set up styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Error/Warning styles for conditional formatting matching HTML theme
        error_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Light red
        warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Light yellow
        error_misconf_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Very light red
        warning_misconf_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Very light yellow
        error_font = Font(color="721C24", bold=True)  # Dark red, bold
        warning_font = Font(color="856404", bold=True)  # Dark yellow, bold
        error_misconf_font = Font(color="721C24")  # Dark red, normal
        warning_misconf_font = Font(color="856404")  # Dark yellow, normal
        
        # Write title
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        table_rows = table_data['table_rows']
        headers = table_data['headers']
        aggregates = table_data['aggregates']
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows grouped by namespace with namespace totals
        from data_gatherer.reporting.common import get_rules_engine
        from openpyxl.comments import Comment
        from itertools import groupby
        
        rules_engine = get_rules_engine()
        
        # Sort table rows by namespace, kind, name, container
        table_rows_sorted = sorted(table_rows, key=lambda r: (r[1], r[0], r[2], r[3]))
        
        # Styles for namespace totals rows
        ns_totals_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ns_totals_font = Font(bold=True)
        
        current_row = 4  # Start after headers
        
        # Process rows grouped by namespace
        for ns, rows_group in groupby(table_rows_sorted, key=lambda r: r[1]):
            rows_list = list(rows_group)
            
            # Write namespace header row
            ns_header_cell = ws.cell(row=current_row, column=1, value=ns or "")
            ns_header_cell.font = Font(bold=True)
            ns_header_cell.border = border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
            current_row += 1
            
            # Write workload rows for this namespace
            for row_data in rows_list:
                # Build row_data dictionary for rules engine
                row_dict = {headers[i]: row_data[i] for i in range(min(len(headers), len(row_data)))}
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.border = border
                    
                    # Apply conditional formatting using rules engine (same as HTML)
                    if col_num <= len(headers):
                        column_name = headers[col_num - 1]
                        context = {
                            'cell_value': str(value),
                            'column_name': column_name,
                            'row_data': row_dict,
                            'report_type': 'container-capacity'
                        }
                        result = rules_engine.evaluate_cell(context)
                        
                        # Apply formatting based on rule result
                        if result.rule_type.value == 'error_miss':
                            cell.fill = error_fill
                            cell.font = error_font
                            if result.message:
                                cell.comment = Comment(result.message, "Data Gatherer")
                        elif result.rule_type.value == 'warning_miss':
                            cell.fill = warning_fill
                            cell.font = warning_font
                            if result.message:
                                cell.comment = Comment(result.message, "Data Gatherer")
                        elif result.rule_type.value == 'error_misconf':
                            cell.fill = error_misconf_fill
                            cell.font = error_misconf_font
                            if result.message:
                                cell.comment = Comment(result.message, "Data Gatherer")
                        elif result.rule_type.value == 'warning_misconf':
                            cell.fill = warning_misconf_fill
                            cell.font = warning_misconf_font
                            if result.message:
                                cell.comment = Comment(result.message, "Data Gatherer")
                    
                    # Right-align numeric columns
                    if col_num >= 6:  # Numeric columns
                        cell.alignment = Alignment(horizontal='right')
                
                current_row += 1
            
            # Calculate and write namespace totals row
            if rows_list:
                ns_main_cpu_raw = sum((int(r[6]) for r in rows_list if r[4] == 'main' and r[6] != ''), 0)
                ns_main_cpu_lim_raw = sum((int(r[7]) for r in rows_list if r[4] == 'main' and r[7] != ''), 0)
                ns_main_mem_raw = sum((int(r[8]) for r in rows_list if r[4] == 'main' and r[8] != ''), 0)
                ns_main_mem_lim_raw = sum((int(r[9]) for r in rows_list if r[4] == 'main' and r[9] != ''), 0)
                ns_main_cpu_total = sum((int(r[10]) for r in rows_list if r[4] == 'main' and r[10] != ''), 0)
                ns_main_cpu_lim_total = sum((int(r[11]) for r in rows_list if r[4] == 'main' and r[11] != ''), 0)
                ns_main_mem_total = sum((int(r[12]) for r in rows_list if r[4] == 'main' and r[12] != ''), 0)
                ns_main_mem_lim_total = sum((int(r[13]) for r in rows_list if r[4] == 'main' and r[13] != ''), 0)
                
                # Shared multiline comments (also used for HTML tooltips) built via helper
                comments = self._namespace_totals_comments(ns or '', {
                    'CPU_req_m': ns_main_cpu_raw,
                    'CPU_lim_m': ns_main_cpu_lim_raw,
                    'Mem_req_Mi': ns_main_mem_raw,
                    'Mem_lim_Mi': ns_main_mem_lim_raw,
                    'CPU_req_m_total': ns_main_cpu_total,
                    'CPU_lim_m_total': ns_main_cpu_lim_total,
                    'Mem_req_Mi_total': ns_main_mem_total,
                    'Mem_lim_Mi_total': ns_main_mem_lim_total,
                })
                
                # Write namespace totals label
                ns_label_cell = ws.cell(row=current_row, column=1, value="Namespace totals")
                ns_label_cell.font = ns_totals_font
                ns_label_cell.fill = ns_totals_fill
                ns_label_cell.border = border
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                
                # Write namespace totals values with individual formula comments on each cell
                totals_values = [
                    (7, ns_main_cpu_raw, comments['CPU_req_m']),
                    (8, ns_main_cpu_lim_raw, comments['CPU_lim_m']),
                    (9, ns_main_mem_raw, comments['Mem_req_Mi']),
                    (10, ns_main_mem_lim_raw, comments['Mem_lim_Mi']),
                    (11, ns_main_cpu_total, comments['CPU_req_m_total']),
                    (12, ns_main_cpu_lim_total, comments['CPU_lim_m_total']),
                    (13, ns_main_mem_total, comments['Mem_req_Mi_total']),
                    (14, ns_main_mem_lim_total, comments['Mem_lim_Mi_total'])
                ]
                
                for col_num, value, comment_text in totals_values:
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.font = ns_totals_font
                    cell.fill = ns_totals_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
                    
                    # Add descriptive comment to each totals cell
                    cell.comment = Comment(comment_text, "Data Gatherer")
                    cell.comment.width = 350
                    cell.comment.height = 200
                
                current_row += 1
        
        # Add global totals section (cluster-wide)
        if table_rows:
            main_cpu_raw = aggregates['main_cpu_raw']
            main_mem_raw = aggregates['main_mem_raw']
            main_cpu_total = aggregates['main_cpu_total']
            main_mem_total = aggregates['main_mem_total']
            main_cpu_lim_raw = aggregates['main_cpu_lim_raw']
            main_mem_lim_raw = aggregates['main_mem_lim_raw']
            main_cpu_lim_total = aggregates['main_cpu_lim_total']
            main_mem_lim_total = aggregates['main_mem_lim_total']
            
            # Add spacing row before global totals
            current_row += 1
            
            # Main containers totals
            ws.cell(row=current_row, column=1, value="Totals (main containers)").font = Font(bold=True)
            ws.cell(row=current_row, column=7, value=main_cpu_raw)
            ws.cell(row=current_row, column=8, value=main_cpu_lim_raw)
            ws.cell(row=current_row, column=9, value=main_mem_raw)
            ws.cell(row=current_row, column=10, value=main_mem_lim_raw)
            ws.cell(row=current_row, column=11, value=main_cpu_total)
            ws.cell(row=current_row, column=12, value=main_cpu_lim_total)
            ws.cell(row=current_row, column=13, value=main_mem_total)
            ws.cell(row=current_row, column=14, value=main_mem_lim_total)
            
            # All containers totals
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(out_path)

    def _namespace_totals_tooltips(self) -> Dict[str, str]:
        # Deprecated concise tooltips retained for backward compatibility (currently unused)
        return {
            'label': 'Namespace totals.',
            'CPU_req_m': 'Sum of per-container CPU requests (m) for main containers (not multiplied by replicas).',
            'CPU_lim_m': 'Sum of per-container CPU limits (m) for main containers (not multiplied by replicas).',
            'Mem_req_Mi': 'Sum of per-container memory requests (Mi) for main containers (not multiplied by replicas).',
            'Mem_lim_Mi': 'Sum of per-container memory limits (Mi) for main containers (not multiplied by replicas).',
            'CPU_req_m_total': 'Sum(replicas × per-container CPU request) for main containers.',
            'CPU_lim_m_total': 'Sum(replicas × per-container CPU limit) for main containers.',
            'Mem_req_Mi_total': 'Sum(replicas × per-container memory request) for main containers.',
            'Mem_lim_Mi_total': 'Sum(replicas × per-container memory limit) for main containers.'
        }

    def _namespace_totals_comments(self, ns: str, values: Dict[str, int]) -> Dict[str, str]:
        """Return unified multi-line comment/tooltip text for namespace totals (HTML + Excel).

        Args:
            ns: Namespace name (may be empty string)
            values: Dict of numeric totals keyed by metric name
        Returns:
            Dict mapping metric names (and 'label') to multi-line explanatory text.
        """
        # Only show MiB, no GiB conversion
        def fmt_mib(mib: int) -> str:
            return f"{mib} MiB"

        comments = {
            'label': f"Namespace totals for '{ns}' (runtime steady-state containers only).",
            'CPU_req_m': (
                f"Sum of CPU requests for all main containers in namespace {ns}.\n"
                "Formula: Σ(CPU_req_m) for each main container (not multiplied by replicas)."
            ),
            'CPU_lim_m': (
                f"Sum of CPU limits for all main containers in namespace {ns}.\n"
                "Formula: Σ(CPU_lim_m) for each main container (not multiplied by replicas)."
            ),
            'Mem_req_Mi': (
                f"Sum of memory requests for all main containers in namespace {ns}.\n"
                "Formula: Σ(Mem_req_Mi) for each main container (not multiplied by replicas)."
            ),
            'Mem_lim_Mi': (
                f"Sum of memory limits for all main containers in namespace {ns}.\n"
                "Formula: Σ(Mem_lim_Mi) for each main container (not multiplied by replicas)."
            ),
            'CPU_req_m_total': (
                f"Total CPU requests for all running pods in namespace {ns}.\n"
                "Formula: Σ(Replicas × CPU_req_m) for each main container."
            ),
            'CPU_lim_m_total': (
                f"Total CPU limits for all running pods in namespace {ns}.\n"
                "Formula: Σ(Replicas × CPU_lim_m) for each main container."
            ),
            'Mem_req_Mi_total': (
                f"Total memory requests for all running pods in namespace {ns}.\n"
                "Formula: Σ(Replicas × Mem_req_Mi) for each main container."
            ),
            'Mem_lim_Mi_total': (
                f"Total memory limits for all running pods in namespace {ns}.\n"
                "Formula: Σ(Replicas × Mem_lim_Mi) for each main container."
            ),
        }
        return comments
