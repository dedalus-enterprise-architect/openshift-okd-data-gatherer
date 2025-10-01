from __future__ import annotations
from .base import ReportGenerator, register
from ..persistence.db import WorkloadDB
from ..persistence.workload_queries import WorkloadQueries
from .common import (
    CONTAINER_WORKLOAD_KINDS, cpu_to_milli, mem_to_mi,
    extract_pod_spec, get_replicas_for_workload,
    build_legend_html, get_common_legend_sections, wrap_html_document,
    format_cell_with_condition
)
import html
import os


@register
class ContainerConfigurationReport(ReportGenerator):
    type_name = 'containers-config'
    file_extension = '.html'
    filename_prefix = 'containers-config-'
    supported_formats = ['html', 'excel']

    def generate(self, db: WorkloadDB, cluster: str, out_path: str, format: str = 'html') -> None:
        # Generate the core data
        table_rows, headers = self._generate_data(db, cluster)
        title = f"Container Configuration Report: {cluster}"
        
        if format.lower() == 'excel':
            self._generate_excel(title, headers, table_rows, out_path)
        else:
            # Default to HTML
            html_content = self._build_html_document(title, headers, table_rows, cluster)
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

    def _generate_data(self, db: WorkloadDB, cluster: str):
        """Generate the core data structure used by both HTML and Excel formats."""
        wq = WorkloadQueries(db)
        rows = wq.list_for_kinds(cluster, list(CONTAINER_WORKLOAD_KINDS))
        table_rows = []
        for rec in rows:
            kind = rec['kind']
            namespace = rec['namespace']
            name = rec['name']
            manifest = rec['manifest']
            pod_spec = extract_pod_spec(kind, manifest)
            if not pod_spec:
                continue
            replicas = get_replicas_for_workload(kind, manifest)
            pod_labels = self._format_labels(
                (manifest.get('spec', {}).get('template', {}).get('metadata', {}).get('labels', {}))
            )
            node_selector = self._format_node_selector(pod_spec.get('nodeSelector', {}))
            containers = [(c, 'main') for c in pod_spec.get('containers', [])]
            containers += [(c, 'init') for c in pod_spec.get('initContainers', [])]
            for cdef, ctype in containers:
                container_name = cdef.get('name', 'Unknown')
                resources = cdef.get('resources', {})
                requests = resources.get('requests', {})
                limits = resources.get('limits', {})
                cpu_req = cpu_to_milli(requests.get('cpu'))
                cpu_lim = cpu_to_milli(limits.get('cpu'))
                mem_req = mem_to_mi(requests.get('memory'))
                mem_lim = mem_to_mi(limits.get('memory'))
                readiness_probe = self._extract_readiness_probe_timeout(cdef)
                image_pull_policy = cdef.get('imagePullPolicy', 'IfNotPresent')
                java_opts = self._extract_java_opts(cdef, namespace, db)
                row = [
                    kind,
                    namespace,
                    name,
                    container_name,
                    ctype,
                    replicas if replicas is not None else '',
                    cpu_req if cpu_req is not None else '',
                    cpu_lim if cpu_lim is not None else '',
                    mem_req if mem_req is not None else '',
                    mem_lim if mem_lim is not None else '',
                    readiness_probe,
                    image_pull_policy,
                    node_selector,
                    pod_labels,
                    java_opts
                ]
                table_rows.append(row)
        
        headers = [
            "Kind", "Namespace", "Name", "Container", "Type",
            "Replicas", "CPU_req_m", "CPU_lim_m", "Mem_req_Mi", "Mem_lim_Mi",
            "Readiness_Probe", "Image_Pull_Policy", "Node_Selectors", "Pod_Labels", "Java_Opts"
        ]
        return table_rows, headers

    def _generate_excel(self, title: str, headers: list, table_rows: list, out_path: str) -> None:
        """Generate Excel output with formatting and conditional highlighting."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Container Configuration"
        
        # Set up styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Error/Warning styles for conditional formatting
        error_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        error_font = Font(color="721C24")
        warning_font = Font(color="856404")
        
        # Write title
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data rows with conditional formatting
        for row_idx, row_data in enumerate(table_rows, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                
                # Apply conditional formatting based on column and value
                header_name = headers[col_idx - 1] if col_idx <= len(headers) else ""
                
                # Apply formatting rules similar to HTML version
                if header_name in ["CPU_req_m", "Mem_req_Mi"] and (value == "" or value is None):
                    cell.fill = error_fill
                    cell.font = error_font
                elif header_name in ["CPU_lim_m", "Mem_lim_Mi"] and (value == "" or value is None):
                    cell.fill = warning_fill
                    cell.font = warning_font
                elif header_name == "Readiness_Probe" and value == "Not configured":
                    cell.fill = error_fill
                    cell.font = error_font
                elif header_name == "Image_Pull_Policy" and value == "Always":
                    cell.fill = warning_fill
                    cell.font = warning_font
                
                # Set alignment based on column type
                if header_name in ["CPU_req_m", "CPU_lim_m", "Mem_req_Mi", "Mem_lim_Mi", "Replicas"]:
                    cell.alignment = Alignment(horizontal="right")
                elif header_name in ["Kind", "Type"]:
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            header_name = headers[col - 1]
            
            # Set specific widths for known columns
            if header_name in ["Kind", "Type"]:
                ws.column_dimensions[column_letter].width = 12
            elif header_name in ["CPU_req_m", "CPU_lim_m", "Mem_req_Mi", "Mem_lim_Mi", "Replicas"]:
                ws.column_dimensions[column_letter].width = 10
            elif header_name in ["Node_Selectors", "Pod_Labels", "Java_Opts"]:
                ws.column_dimensions[column_letter].width = 40
            elif header_name in ["Namespace", "Name", "Container"]:
                ws.column_dimensions[column_letter].width = 20
            else:
                ws.column_dimensions[column_letter].width = 15
        
        # Add summary row
        summary_row = len(table_rows) + 5
        ws.cell(row=summary_row, column=1, value="Total containers:")
        ws.cell(row=summary_row, column=2, value=len(table_rows))
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=2).font = Font(bold=True)
        
        # Save the workbook
        wb.save(out_path)

    def _extract_readiness_probe_timeout(self, container_def):
        readiness_probe = container_def.get('readinessProbe', {})
        if readiness_probe:
            timeout = readiness_probe.get('timeoutSeconds', 1)
            initial_delay = readiness_probe.get('initialDelaySeconds', 0)
            return f"{timeout}s (initial: {initial_delay}s)"
        return "Not configured"

    def _extract_java_opts(self, container_def, namespace: str, db: WorkloadDB):
        # Direct env value first
        env = container_def.get('env', [])
        for env_var in env:
            var_name = env_var.get('name', '').upper()
            if 'JAVA' in var_name and 'OPT' in var_name:
                value = env_var.get('value', '')
                if value:
                    return value
        # valueFrom -> configMapKeyRef
        for env_var in env:
            value_from = env_var.get('valueFrom', {})
            cm_ref = value_from.get('configMapKeyRef') if isinstance(value_from, dict) else None
            if not cm_ref:
                continue
            key_name = env_var.get('name', '').upper()
            if 'JAVA' in key_name and 'OPT' in key_name:
                cm_name = cm_ref.get('name')
                cm_key = cm_ref.get('key')
                val = self._lookup_configmap_value(db, namespace, cm_name, cm_key)
                if val:
                    return val
        # envFrom configMapRef entire data scan for likely JAVA options
        for env_from in container_def.get('envFrom', []) or []:
            cm_ref = env_from.get('configMapRef') if isinstance(env_from, dict) else None
            if not cm_ref:
                continue
            cm_name = cm_ref.get('name')
            data = self._lookup_configmap_data(db, namespace, cm_name)
            if data:
                # look for keys containing JAVA_OPTS or similar
                for k, v in data.items():
                    ku = k.upper()
                    if 'JAVA' in ku and 'OPT' in ku and v:
                        return v
        return "Not configured"

    def _lookup_configmap_value(self, db: WorkloadDB, namespace: str, name: str, key: str):
        if not (name and key):
            return None
        cur = db._conn.cursor()
        row = cur.execute(
            "SELECT manifest_json FROM workload WHERE kind=? AND namespace=? AND name=? LIMIT 1",
            ('ConfigMap', namespace, name)
        ).fetchone()
        if not row:
            return None
        import json
        try:
            manifest = json.loads(row[0])
            return (manifest.get('data') or {}).get(key)
        except Exception:
            return None

    def _lookup_configmap_data(self, db: WorkloadDB, namespace: str, name: str):
        if not name:
            return None
        cur = db._conn.cursor()
        row = cur.execute(
            "SELECT manifest_json FROM workload WHERE kind=? AND namespace=? AND name=? LIMIT 1",
            ('ConfigMap', namespace, name)
        ).fetchone()
        if not row:
            return None
        import json
        try:
            manifest = json.loads(row[0])
            return manifest.get('data') or {}
        except Exception:
            return None

    def _format_labels(self, labels_dict):
        if not labels_dict:
            return "None"
        formatted_labels = [f"{k}={v}" for k, v in labels_dict.items()]
        return ", ".join(formatted_labels)

    def _format_node_selector(self, node_selector_dict):
        if not node_selector_dict:
            return "None"
        selectors = [f"{k}={v}" for k, v in node_selector_dict.items()]
        return ", ".join(selectors)

    def _build_html_document(self, title, headers, table_rows, cluster):
        parts = []
        parts.append(f'<h1>{title}</h1>')
        parts.append('<p>Complete container configuration analysis including resource allocation, health settings, and deployment configuration.</p>')
        # Comprehensive column legend: every table header must be represented here
        legend_sections = get_common_legend_sections() + [
            {
                "title": "Key Columns",
                "items": [
                    "<strong>Kind</strong>: Workload controller kind (Deployment / StatefulSet / etc.)",
                    "<strong>Namespace</strong>: Kubernetes namespace (blank if cluster‑scoped)",
                    "<strong>Name</strong>: Workload name",
                    "<strong>Container</strong>: Container name inside spec",
                    "<strong>Type</strong>: main or init",
                    "<strong>Replicas</strong>: Desired replicas (DaemonSet blank)",
                    "<strong>CPU_req_m</strong>: CPU requests in millicores",
                    "<strong>CPU_lim_m</strong>: CPU limits in millicores",
                    "<strong>Mem_req_Mi</strong>: Memory requests in MiB",
                    "<strong>Mem_lim_Mi</strong>: Memory limits in MiB",
                    "<strong>Readiness_Probe</strong>: timeout / initial delay or Not configured",
                    "<strong>Image_Pull_Policy</strong>: Container image pull policy",
                    "<strong>Node_Selectors</strong>: nodeSelector key=value list or None",
                    "<strong>Pod_Labels</strong>: Pod template labels key=value list or None",
                    "<strong>Java_Opts</strong>: Discovered JAVA options (env / ConfigMap) or Not configured"
                ]
            }
        ]
        legend_html = build_legend_html(legend_sections)
        parts.append(legend_html)
        if not table_rows:
            parts.append('<p>No container workloads found.</p>')
        else:
            parts.append(f'<p><strong>Total containers:</strong> {len(table_rows)}</p>')
            parts.append('<table class="report-table">')
            parts.append('<thead><tr>')
            for header in headers:
                parts.append(f'<th>{html.escape(header)}</th>')
            parts.append('</tr></thead>')
            parts.append('<tbody>')
            for row in table_rows:
                parts.append('<tr>')
                for i, cell in enumerate(row):
                    cell_str = str(cell) if cell is not None else ''
                    if i < len(headers):
                        row_data = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
                        parts.append(format_cell_with_condition(cell_str, headers[i], row_data, 'containers'))
                    else:
                        parts.append(f'<td>{html.escape(cell_str)}</td>')
                parts.append('</tr>')
            parts.append('</tbody>')
            parts.append('</table>')
        return wrap_html_document(title, parts, self._get_unified_containers_css())

    def _get_unified_containers_css(self):
        return """
        .report-table { font-size: 12px; }
        .report-table th { position: sticky; top: 0; z-index: 10; }
        .report-table td { word-wrap: break-word; word-break: break-word; white-space: normal; }
        .report-table td:nth-child(1) { font-weight: bold; }
        .report-table td:nth-child(2) { font-family: monospace; }
        .report-table td:nth-child(3) { font-weight: 500; }
        .report-table td:nth-child(4) { font-family: monospace; }
        .report-table td:nth-child(5) { font-weight: bold; text-align: center; }
        .report-table td:nth-child(7),
        .report-table td:nth-child(8),
        .report-table td:nth-child(9),
        .report-table td:nth-child(10) { font-family: monospace; text-align: right; }
        .report-table td:nth-child(11) { font-size: 11px; }
        .report-table td:nth-child(12) { font-weight: 500; }
        .report-table td:nth-child(13) { font-size: 11px; }
        .report-table td:nth-child(14) { font-size: 10px; min-width: 200px; }
        .report-table td:nth-child(15) { font-family: monospace; font-size: 10px; min-width: 250px; }
        @media (max-width: 1400px) { .report-table { font-size: 11px; } .report-table td { padding: 6px; } .report-table td:nth-child(14), .report-table td:nth-child(15) { min-width: 150px; } }
        @media (max-width: 900px) { .report-table { display: block; overflow-x: auto; white-space: nowrap; } }
        </style>
        """
